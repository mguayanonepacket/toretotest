# Jose Carbonell
# v0.3
from NBFunctions import *
from postchecks import getNokiaInformation
import ipaddress
import sys
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, PatternFill, Border, Side, Alignment
from copy import copy
import pprint
import subprocess
from paramiko import SSHClient, AutoAddPolicy
import pprint
import re
from lxml import etree


def checkGSVersion(conn, host, goldStandardVersion, deviceType):
    # Function that gets the code version from the devices and compares it againts the GS standard version that it is in this wiki
    # https://packet.atlassian.net/wiki/spaces/OP/pages/2442527156/NETOPS-SOP+NOS+Gold+Master+Versions. Retrurs True if the test is successful and
    # False if firmware versions don't match. It also print the errors in the REPORT is the firmaware versions don't match.
    commandsReport = {}
    errors = {}
    if deviceType == "arista_eos":  # Arista
        command = "show version"
        i = 1
        try:
            output = conn.enable(command)

        except:
            print(
                "\nWARNING: The command "
                + command
                + " failed for the device "
                + host
                + ". Check the command syntax or run it manually on the device. Skipping Test 1 ...\n"
            )
            return

        currentCodeVersion = output[0]["result"]["version"]

    if deviceType == "juniper_junos":  # Juniper
        command = "show version"
        i = 1
        try:
            facts = conn.facts

        except:
            print(
                "\nWARNING: The command "
                + command
                + " failed for the device "
                + host
                + ". Check the command syntax or run it manually on the device. Skipping Test 1 ...\n"
            )
            return

        currentCodeVersion = (
            conn.facts["version"].split(".")[0]
            + "."
            + conn.facts["version"].split(".")[1]
            + ("." + conn.facts["version"].split(".")[2] if len(conn.facts["version"].split(".")) > 2 else "")
        )

    if deviceType == "sr_linux":  # Nokia
        command = "platform"
        i = 1
        try:

            result = getNokiaInformation(host, "platform")

        except:
            print(
                "\nWARNING: The command "
                + command
                + " failed for the device "
                + host
                + ". Check the command syntax or run it manually on the device. Skipping Test 1 ...\n"
            )
            return

        currentCodeVersion = result["notification"][0]["update"][0]["val"]["srl_nokia-platform-control:control"][0][
            "software-version"
        ].split("-")[0]

    if not (currentCodeVersion == goldStandardVersion):

        # print("Test 1 failed: The device "+host+" is not in Gold Standard firmware version. Its current firmware version is: "+conn.facts['version'])
        errors[i] = (
            "The device "
            + host
            + " is not in Gold Standard firmware version. Its current firmware version is: "
            + currentCodeVersion
            + " - The Gold Standard Version is "
            + goldStandardVersion
        )
        commandsReport[i] = {1: command}
        printReport(
            conn,
            host,
            "Test 1 failed for " + host + " - The device " + host + " is not in Gold Standard firmware version: \n",
            errors,
            commandsReport,
            deviceType,
        )
        i = i + 1
        return False
    else:
        # print("Test 1 successfull: The device "+host+" is in Gold Standard firmware version: "+output[0]['result']['version'])
        printReport(
            conn,
            host,
            "Test 1 successfull: The device " + host + " is in Gold Standard firmware version: " + currentCodeVersion,
            errors,
            commandsReport,
            deviceType,
        )
        return True


def checkDevSNNB(conn, host, deviceType):
    # Function that checks the device sn and compares it againts the sn that is in netbox for the device. Retrurs True if the test is successful and
    # False if sns don't match. It also print the errors in the REPORT in case of mismatch.

    NBserialNumber = getSNfromNB(host)
    commandsReport = {}
    errors = {}
    devSN = {}
    i = 1
    if deviceType == "arista_eos":  # Arista
        command = "show version"
        try:
            output = conn.enable(command)

        except:
            print(
                "\nWARNING: The command "
                + command
                + " failed for the device "
                + host
                + ". Check the command syntax or run it manually on the device. Skipping Test 2 ...\n"
            )
            return

        devSN = {"sn": output[0]["result"]["serialNumber"], "command": command}

    if deviceType == "juniper_junos":  # Juniper
        command = "show chassis hardware"
        try:
            facts = conn.facts

        except:
            print(
                "\nWARNING: The serial Number for the device "
                + host
                + " couldn't be obtanied. Check the command syntax or run it manually on the device. Skipping Test 2 ...\n"
            )
            return

        devSN = {"sn": facts["serialnumber"], "command": command}

    if deviceType == "sr_linux":  # Nokia
        command = "platform"
        i = 1
        try:

            result = getNokiaInformation(host, "platform")

        except:
            print(
                "\nWARNING: The command "
                + command
                + " failed for the device "
                + host
                + ". Check the command syntax or run it manually on the device. Skipping Test 2 ...\n"
            )
            return

        devSN = {
            "sn": result["notification"][0]["update"][0]["val"]["srl_nokia-platform-control:control"][0][
                "serial-number"
            ],
            "command": command,
        }

    if not (devSN["sn"] == NBserialNumber):

        errors[i] = (
            "The device "
            + host
            + " SN does not match the SN created in Netbox. The correct Netbox SN for the device should be: "
            + devSN["sn"]
        )
        commandsReport[i] = {1: devSN["command"]}
        printReport(
            conn,
            host,
            "Test 2 failed for "
            + host
            + " - The device SN does not match the SN created in Netbox for that device: \n",
            errors,
            commandsReport,
            deviceType,
        )
        return False
    else:
        # print("Test 2 successfull: The device "+host+" SN matches the SN created in Netbox: "+devSN['sn'])
        printReport(
            conn,
            host,
            "Test 2 successfull: The device " + host + " SN matches the SN created in Netbox: " + devSN["sn"],
            "",
            "",
            deviceType,
        )
        return True


def checkDevInterfacesNB(conn, host, deviceType, siteName):
    # Function that compares the device interfaces/connections/neighbors (derived from lldp) with interfaces/connections/neighbors that the device has in Netbox.
    # Returns True if the test is successful and False if interfaces/connections/neighbors don't match. It also print the errors in the REPORT in case of mismatch.

    interfacesNB = getDevInterfacesNB(host)
    commandsReport = {}

    if deviceType == "arista_eos":  # Arista

        command = "show lldp neighbors"
        try:
            output = conn.enable(command)

        except:
            print(
                "\nWARNING: The command "
                + command
                + " failed for the device "
                + host
                + ". Check the command syntax or run it manually on the device. Skipping Test 3 ...\n"
            )
            return

        n = 1
        lldpNeigInfo = {}
        for lldpNeighbors in output[0]["result"]["lldpNeighbors"]:
            if (
                lldpNeighbors["neighborDevice"].find(siteName) != -1
            ):  # If the neighbor name contains the site name (to remove non-network devices from lldp neighbor list)
                lldpNeigInfo[n] = {
                    "localHost": host,
                    "localInterface": lldpNeighbors["port"],
                    "neighbor": lldpNeighbors["neighborDevice"],
                    "neighborInterface": lldpNeighbors["neighborPort"],
                }
                n = n + 1

    if deviceType == "juniper_junos":  # Juniper
        lldpNeigInfo = {}
        n = 1
        for n in interfacesNB.keys():
            if (
                interfacesNB[n]["localInterface"].find("fxp0") == -1 and not interfacesNB[n]["mode"] == "tagged"
            ):  # Skip subinterfaces ECX CORE/Fabric and fxp0
                command = "show lldp neighbors interface " + str(interfacesNB[n]["localInterface"])
                try:
                    lldpNeighbors_info = conn.rpc.get_lldp_interface_neighbors(
                        interface_device=interfacesNB[n]["localInterface"]
                    )

                except:
                    print(
                        "\nWARNING:The lldp neighbors for the device "
                        + host
                        + " couldn't be obtanied for the interface "
                        + str(interfacesNB[n]["localInterface"])
                        + " . Check the command syntax or run it manually on the device. Skipping Test 3 JUN...\n"
                    )
                    continue

                lldpNeigList = lldpNeighbors_info.findall(".//lldp-neighbor-information")
                for lldpNeighbors in lldpNeigList:
                    if (
                        lldpNeighbors.find(".//lldp-remote-system-name").text.find(siteName) != -1
                    ):  # IF THE NEIGHBOR NAME CONTAINS THE SITE NAME (TO REMOVE NON-NETWORK DEVICES FROM LLDP NEIGHBOR LIST)
                        lldpNeigInfo[n] = {
                            "localHost": host,
                            "localInterface": lldpNeighbors.find(".//lldp-local-interface").text,
                            "neighbor": lldpNeighbors.find(".//lldp-remote-system-name").text,
                            "neighborInterface": lldpNeighbors.find(".//lldp-remote-port-id").text,
                        }
                        n = n + 1

    if deviceType == "sr_linux":  # Nokia
        lldpNeigInfo = {}
        command = "system/lldp/interface"
        n = 1
        try:

            result = getNokiaInformation(host, "system/lldp/interface")

        except:
            print(
                "\nWARNING: The command "
                + command
                + " failed for the device "
                + host
                + ". Check the command syntax or run it manually on the device. Skipping Test 3 ...\n"
            )
            return

        for interfaces in result["notification"][0]["update"][0]["val"]["interface"]:
            if "neighbor" in interfaces.keys() and interfaces["neighbor"][0]["system-name"]:
                lldpNeigInfo[n] = {
                    "localHost": host,
                    "localInterface": interfaces["name"],
                    "neighbor": interfaces["neighbor"][0]["system-name"],
                    "neighborInterface": interfaces["neighbor"][0]["port-id"],
                }
                n = n + 1

    differences = compareInterfaces(lldpNeigInfo, interfacesNB)
    if differences:

        for i in differences.keys():
            if deviceType == "sr_linux":
                commandsReport[i] = {1: str(command) + "[name=" + str(differences[i]["localInterface"]) + "]"}
            else:
                commandsReport[i] = {1: command, 2: "show interfaces descriptions"}

        printReport(
            conn,
            host,
            "Test 3 failed for "
            + host
            + " - Found one or more interfaces where lldp information doesn't match the Netbox information for the device: \n",
            differences,
            commandsReport,
            deviceType,
        )
        return False

    else:

        # print("Test 3 successfull for "+host+" - the netbox interfaces match the lldp neighbor interfaces")
        printReport(
            conn,
            host,
            "Test 3 successfull for " + host + " - the netbox interfaces match the lldp neighbor interfaces",
            "",
            "",
            deviceType,
        )
        return True


def getInterfaceAdapterJuniper(interfaceInfo):
    # Function required to adapt the arista response dictionary to standard dictionary format (not vendor specific) so that the code of the checkInterfacesStatus function can be used for both vendors
    # Returns a dictionary with all the interface status information in standard format

    return_dict = {}
    for interfaces in interfaceInfo:
        interfaceStatus = (
            interfaces.find(".//admin-status").text.strip("\n")
            if interfaces.find(".//admin-status") is not None
            else "Down"
        )
        lineProtocolStatus = (
            interfaces.find(".//oper-status").text.strip("\n")
            if interfaces.find(".//oper-status") is not None
            else "Down"
        )
        return_dict[interfaces.find(".//name").text.strip("\n")] = {
            "interfaceStatus": interfaceStatus,
            "lineProtocolStatus": lineProtocolStatus,
            "description": interfaces.find(".//description").text.strip("\n"),
            "name": interfaces.find(".//name").text.strip("\n"),
        }

    return return_dict


def getInterfaceAdapterArista(interfaceInfo):
    # Function required to adapt the Arista response dictionary to standard dictionary format (not vendor specific) so that the code of the checkInterfacesStatus function can be used for both vendors
    # Returns a dictionary with all the interface status information in standard format

    return_dict = {}
    for interfaces in interfaceInfo.keys():

        return_dict[interfaces] = {
            "interfaceStatus": interfaceInfo[interfaces]["interfaceStatus"],
            "lineProtocolStatus": interfaceInfo[interfaces]["lineProtocolStatus"],
            "description": interfaceInfo[interfaces]["description"],
            "name": interfaces,
        }

    return return_dict


def getInterfaceAdapterNokia(interfaceInfo):
    # Function required to adapt the Arista response dictionary to standard dictionary format (not vendor specific) so that the code of the checkInterfacesStatus function can be used for both vendors
    # Returns a dictionary with all the interface status information in standard format

    return_dict = {}
    for interfaces in interfaceInfo["notification"][0]["update"][0]["val"]["srl_nokia-interfaces:interface"]:
        if "description" in interfaces.keys():
            return_dict[interfaces["name"]] = {
                "interfaceStatus": interfaces["oper-state"],
                "lineProtocolStatus": interfaces["admin-state"],
                "description": interfaces["description"],
                "name": interfaces["name"],
            }
        else:
            return_dict[interfaces["name"]] = {
                "interfaceStatus": interfaces["oper-state"],
                "lineProtocolStatus": interfaces["admin-state"],
                "description": "",
                "name": interfaces["name"],
            }
    return return_dict


def checkInterfacesStatus(conn, host, deviceType):
    # Function that checks the status of the interfaces (both admin and protocol status).
    # Returns True if the test is successful and False if it founds interfaces in down state. It also print the errors in the REPORT in case of interfaces down.

    interfacesNB = getDevInterfacesNB(host)
    errors = {}
    commandsReport = {}
    i = 1
    if deviceType == "arista_eos":  # Arista

        command = "show interfaces description"
        try:
            output = conn.enable(command)

        except:
            print(
                "\nWARNING: The command '"
                + command
                + "' failed for the device "
                + host
                + ". Check the command syntax or run it manually on the device. Skipping Test 4 ...\n"
            )
            return

        interfaceList = getInterfaceAdapterArista(output[0]["result"]["interfaceDescriptions"])

    if deviceType == "juniper_junos":  # Juniper
        command = "show interfaces descriptions"
        try:
            interfaceInfo = conn.rpc.get_interface_information(descriptions=True)

        except:
            print(
                "\nWARNING: The command 'show interface descriptions' failed for the device "
                + host
                + ". Check the command syntax or run it manually on the device. Skipping Test 4 ...\n"
            )
            return

        interfaceList = getInterfaceAdapterJuniper(interfaceInfo.findall(".//physical-interface"))
        interfaceList.update(getInterfaceAdapterJuniper(interfaceInfo.findall(".//logical-interface")))

    if deviceType == "sr_linux":  # Nokia
        command = "interface"
        n = 1
        try:
            result = getNokiaInformation(host, "interface")

        except:
            print(
                "\nWARNING: The command "
                + command
                + " failed for the device "
                + host
                + ". Check the command syntax or run it manually on the device. Skipping Test 3 ...\n"
            )
            return

        interfaceList = getInterfaceAdapterNokia(result)

    for n in interfacesNB.keys():

        if interfacesNB[n]["localInterface"] in interfaceList.keys():

            if interfaceList[interfacesNB[n]["localInterface"]]["interfaceStatus"] != "up":
                errors[i] = (
                    host + " - The interface " + interfacesNB[n]["localInterface"] + " has interface status down"
                )
                if deviceType == "sr_linux":
                    commandsReport[i] = {1: str(command) + "[name=" + str(interfacesNB[n]["localInterface"]) + "]"}
                    i = i + 1

                else:
                    commandsReport[i] = {1: command}
                    i = i + 1

            elif (
                interfaceList[interfacesNB[n]["localInterface"]]["lineProtocolStatus"] != "up"
                and interfaceList[interfacesNB[n]["localInterface"]]["lineProtocolStatus"] != "enable"
            ):
                errors[i] = host + " - The interface " + interfacesNB[n]["localInterface"] + " has protocol status down"
                if deviceType == "sr_linux":
                    commandsReport[i] = {1: str(command) + "[name=" + str(interfacesNB[n]["localInterface"]) + "]"}
                    i = i + 1
                else:
                    commandsReport[i] = {1: command}
                    i = i + 1

    if errors:

        printReport(
            conn,
            host,
            "Test 4 failed for " + host + " - Found one or more interfaces in down state: \n",
            errors,
            commandsReport,
            deviceType,
        )
        return False

    else:
        printReport(conn, host, "Test 4 successfull for " + host + " - All interfaces are UP", "", "", deviceType)
        return True


def getIntPWLvAdapterJuniper(interfaceInfo, interfaceName, command):
    # Function required to adapt the Juniper response dictionary to standard dictionary format (not vendor specific), so that the code of the checkIntOptPWLevels function can be used for both vendors
    # Returns a dictionary with all the interface power levels in standard format
    return_dict = {}

    return_dict[interfaceName] = dict(
        (
            lane.find(".//lane-index").text.strip("\n"),
            {
                "txPower": lane.find(".//laser-output-power-dbm").text.strip("\n"),
                "rxPower": lane.find(".//laser-rx-optical-power-dbm").text.strip("\n"),
                "command": command,
            },
        )
        for lane in interfaceInfo
    )

    return return_dict


def getIntPWLvAdapterArista(interfaceInfo, interfaceName, command):
    # Function required to adapt the Arista response dictionary to standard dictionary format (not vendor specific), so that the code of the checkIntOptPWLevels function can be used for both vendors
    # Returns a dictionary with all the interface power levels in standard format
    return_dict = {}
    for interfaces in interfaceInfo.keys():
        return_dict[interfaceName] = {
            0: {
                "txPower": interfaceInfo[interfaces]["txPower"],
                "rxPower": interfaceInfo[interfaces]["rxPower"],
                "command": command,
            }
        }

    return return_dict


def getIntPWLvAdapterNokia(interfaceInfo, interfaceName, command):
    # Function required to adapt the Nokia response dictionary to standard dictionary format (not vendor specific), so that the code of the checkIntOptPWLevels function can be used for both vendors
    # Returns a dictionary with all the interface power levels in standard format
    return_dict = {}
    return_dict[interfaceName] = dict(
        (
            lane["index"],
            {
                "txPower": lane["output-power"]["latest-value"],
                "rxPower": lane["input-power"]["latest-value"],
                "command": command,
            },
        )
        for lane in interfaceInfo["notification"][0]["update"][0]["val"]["transceiver"]["channel"]
    )

    return return_dict


def checkIntOptPWLevels(conn, host, deviceType):
    # Function that checks if interface optic light levels are between +4 and -4 dbms.
    # Returns True if the test is successful and False if it founds interfaces with power levels outside that range. It also print the errors in the REPORT in
    # case of low/high light levels founds.

    interfacesNB = getDevInterfacesNB(host)
    errors = {}
    commandsReport = {}
    interfacePWLevelsList = {}
    i = 1
    lowThreshold = -5  # low threshold in dbm based on experience for simplicity
    highThreshold = 4  # high thrshold in dbm based on experience for simplicity

    if deviceType == "arista_eos":  # Arista
        for n in interfacesNB.keys():
            txErr = ""
            rxErr = ""
            command = "show interfaces " + str(interfacesNB[n]["localInterface"]) + " transceiver detail"
            # if not mgmt interface and not a pdu port (no optic) and interface with an optic
            if (
                interfacesNB[n]["localInterface"].find("Management1") == -1
                and interfacesNB[n]["neighborInterface"].find("eth0") == -1
                and interfacesNB[n]["neighborInterface"].find("Management1") == -1
            ):
                try:
                    output = conn.enable(command)

                except:
                    print(
                        "\nWARNING: The command '"
                        + command
                        + "' failed for the device "
                        + host
                        + ". Check the command syntax or run it manually on the device. Skipping Test 5 ...\n"
                    )
                    continue
                if len(output[0]["result"]["interfaces"]) > 0:
                    if (
                        len(output[0]["result"]["interfaces"][interfacesNB[n]["localInterface"]]) > 0
                    ):  # this is done because sometimes dac cables are used for some links and they don't report any power level

                        interfacePWLevelsList.update(
                            getIntPWLvAdapterArista(
                                output[0]["result"]["interfaces"], interfacesNB[n]["localInterface"], command
                            )
                        )

    if deviceType == "juniper_junos":  # Juniper

        for n in interfacesNB.keys():
            txErr = ""
            rxErr = ""
            command = "show interfaces diagnostics optics " + str(interfacesNB[n]["localInterface"])
            if interfacesNB[n]["localInterface"].find("fxp0") == -1 and (
                interfacesNB[n]["neighborInterface"].find("eth0") == -1
                or interfacesNB[n]["neighborInterface"].find("eth1") == -1
            ):  # if not mgmt interface and not a pdu port (no optic), console ports and interface with an optic
                try:
                    interfaceInfo = conn.rpc.get_interface_optics_diagnostics_information(
                        interface_name=str(interfacesNB[n]["localInterface"])
                    )

                except:
                    print(
                        "\nWARNING: The command 'show interface descriptions' failed for the device "
                        + host
                        + ". Check the command syntax or run it manually on the device. Skipping Test 5 ...\n"
                    )
                    continue

                if len(interfaceInfo) > 0:

                    interfacePWLevelsList.update(
                        getIntPWLvAdapterJuniper(
                            interfaceInfo.findall(".//optics-diagnostics-lane-values"),
                            interfacesNB[n]["localInterface"],
                            command,
                        )
                    )

    if deviceType == "sr_linux":  # Nokia
        for n in interfacesNB.keys():
            command = "interface[name=" + str(interfacesNB[n]["localInterface"]) + "]"
            if (
                interfacesNB[n]["localInterface"].find("mgmt") == -1
                and interfacesNB[n]["neighborInterface"].find("system") == -1
            ):  # if not mgmt interface and not a pdu port (no optic), console ports and interface with an optic
                try:
                    interfaceInfo = getNokiaInformation(host, command)

                except:
                    print(
                        "\nWARNING: The command "
                        + command
                        + " failed for the device "
                        + host
                        + ". Check the command syntax or run it manually on the device. Skipping Test 3 ...\n"
                    )
                    return

                if len(interfaceInfo) > 0 and interfaceInfo["notification"][0]["update"][0]["val"]["transceiver"].get(
                    "ethernet-pmd"
                ):
                    if (
                        interfaceInfo["notification"][0]["update"][0]["val"]["transceiver"]["ethernet-pmd"].find(
                            "100GBASE-CR4"
                        )
                        == -1
                    ):  # skip dac cables as they don't report power levels

                        interfacePWLevelsList.update(
                            getIntPWLvAdapterNokia(interfaceInfo, interfacesNB[n]["localInterface"], command)
                        )

    for n in interfacesNB.keys():
        if (
            not interfacesNB[n]["mode"] == "tagged"
            and interfacesNB[n]["localInterface"] in interfacePWLevelsList.keys()
        ):
            for intLanes in interfacePWLevelsList[interfacesNB[n]["localInterface"]].keys():
                txErr = ""
                rxErr = ""
                if (
                    interfacePWLevelsList[interfacesNB[n]["localInterface"]][intLanes]["txPower"] != "- Inf"
                    and interfacePWLevelsList[interfacesNB[n]["localInterface"]][intLanes]["rxPower"] != "- Inf"
                ):  # IF power value returned
                    if not (
                        float(lowThreshold)
                        < float(interfacePWLevelsList[interfacesNB[n]["localInterface"]][intLanes]["txPower"])
                        < float(highThreshold)
                    ):
                        txErr = (
                            "Tx Power issue for "
                            + host
                            + " interface "
                            + str(interfacesNB[n]["localInterface"])
                            + " - Lane "
                            + str(intLanes)
                            + " - Current TX power value: "
                            + str(interfacePWLevelsList[interfacesNB[n]["localInterface"]][intLanes]["txPower"])
                            + " dbm"
                            + " \n Thresholds: Low TX power warning: "
                            + str(lowThreshold)
                            + " dbm - High TX power warning: "
                            + str(highThreshold)
                            + "dbm"
                        )

                    if not (
                        float(lowThreshold)
                        < float(interfacePWLevelsList[interfacesNB[n]["localInterface"]][intLanes]["rxPower"])
                        < float(highThreshold)
                    ):
                        rxErr = (
                            "Rx Power issue for "
                            + host
                            + " interface "
                            + str(interfacesNB[n]["localInterface"])
                            + " - Lane "
                            + str(intLanes)
                            + " - Current RX power value: "
                            + str(interfacePWLevelsList[interfacesNB[n]["localInterface"]][intLanes]["rxPower"])
                            + " dbm"
                            + " \n Thresholds: Low RX power warning: "
                            + str(lowThreshold)
                            + " dbm - High RX power warning: "
                            + str(highThreshold)
                            + " dbm"
                        )

                    if txErr or rxErr:
                        errors[i] = {"TX": txErr, "RX": rxErr, "Lane": intLanes}
                        commandsReport[i] = {
                            1: interfacePWLevelsList[interfacesNB[n]["localInterface"]][intLanes]["command"]
                        }
                        i = i + 1

                else:  # No power value returned (returned -Inf)

                    if interfacePWLevelsList[interfacesNB[n]["localInterface"]][intLanes]["txPower"] == "- Inf":
                        txErr = (
                            "Tx Power issue for "
                            + host
                            + " interface "
                            + str(interfacesNB[n]["localInterface"])
                            + " - Lane "
                            + str(intLanes)
                            + " - Current TX power value: "
                            + str(interfacePWLevelsList[interfacesNB[n]["localInterface"]][intLanes]["txPower"])
                            + " dbm"
                            + " \n Thresholds: Low TX power warning: "
                            + str(lowThreshold)
                            + " dbm - High TX power warning: "
                            + str(highThreshold)
                            + "dbm"
                        )

                    if interfacePWLevelsList[interfacesNB[n]["localInterface"]][intLanes]["rxPower"] == "- Inf":
                        rxErr = (
                            "Rx Power issue for "
                            + host
                            + " interface "
                            + str(interfacesNB[n]["localInterface"])
                            + " - Lane "
                            + str(intLanes)
                            + " - Current RX power value: "
                            + str(interfacePWLevelsList[interfacesNB[n]["localInterface"]][intLanes]["rxPower"])
                            + " dbm"
                            + " \n Thresholds: Low RX power warning: "
                            + str(lowThreshold)
                            + " dbm - High RX power warning: "
                            + str(highThreshold)
                            + " dbm"
                        )

                    if txErr or rxErr:
                        errors[i] = {"TX": txErr, "RX": rxErr, "Lane": intLanes}
                        commandsReport[i] = {
                            1: interfacePWLevelsList[interfacesNB[n]["localInterface"]][intLanes]["command"]
                        }
                        i = i + 1
    if errors:

        printReport(
            conn,
            host,
            "Test 5 failed for "
            + host
            + " - Found one or more interfaces with rx power or tx power levels outside the power threshold levels: \n",
            errors,
            commandsReport,
            deviceType,
        )
        return False

    else:

        printReport(
            conn,
            host,
            "Test 5 successfull for " + host + " - All the interfaces power levels are within the threshold levels",
            "",
            "",
            deviceType,
        )
        return True


def getDeviceOutput(host, command):
    # Gets the output of the command passed as an argument and returns the output of the command

    ssh_command = ["ssh", '-o "UserKnownHostsFile=/dev/null" -o "StrictHostKeyChecking=no"', "-T", host, command]
    try:
        process = subprocess.Popen(ssh_command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, universal_newlines=True)
        output = ""
        while True:
            line = process.stdout.readline()
            if not line:
                break
            output += line
        return output
    except subprocess.CalledProcessError as e:
        print(
            f"\nWARNING: Connection to the device {host} failed while trying to get the output of the command {command}."
        )
        print("Check the command and try again later.\n")
        return ""


def checkDevLogs(conn, host, deviceType):
    # Function that checks if there are device interfaces flapping in the logs
    # Returns True if the test is successful and False if it founds interfaces flapping. It also print the errors in the REPORT in case flapping interfaces found.

    interfacesNB = getDevInterfacesNB(host)
    logs = {}
    errors = {}
    commandsReport = {}
    i = 1
    if deviceType == "arista_eos":  # Arista

        command = "show logging 100"

        try:
            output = conn.enable(command)

        except:
            print(
                "\nWARNING: The command "
                + command
                + " failed for the device "
                + host
                + ". Check the command syntax or run it manually on the device. Skipping Test 6 ...\n"
            )
            return

        logs = str(output[0]["result"]).splitlines()

    if deviceType == "juniper_junos":  # Juniper

        command = "show log messages | last 150"
        try:
            logsInfo = conn.rpc.get_log(filename="messages")

        except:
            print(
                "\nWARNING: The command "
                + command
                + " failed for the device "
                + host
                + ". Check the command syntax or run it manually on the device. Skipping Test 6 ...\n"
            )
            return

        logs = logsInfo.text.splitlines()

    if deviceType == "sr_linux":  # Nokia

        command = "show system logging file messages | tail -n 200"
        try:
            logsInfo = getDeviceOutput(host, command)

        except:
            print(
                "\nWARNING: The command "
                + command
                + " failed for the device "
                + host
                + ". Check the command syntax or run it manually on the device. Skipping Test 3 ...\n"
            )
            return

        logs = logsInfo.splitlines()

    for n in interfacesNB.keys():
        eventNumbr = 0
        if deviceType == "juniper_junos":
            command = "show log messages | last 150 | match " + str(interfacesNB[n]["localInterface"])

        for line in logs[-150:]:  # Check the last 150 log entries
            if (
                (interfacesNB[n]["localInterface"] in line)
                and ("ACCOUNTING-6-CMD" not in line)
                and ("down" in line or "up" in line)
            ):

                eventNumbr = eventNumbr + 1

                if eventNumbr > 10:
                    errors[i] = {"localInterface": interfacesNB[n]["localInterface"], "logMessage": line}
                    commandsReport[i] = {1: command}
                    i = i + 1
                    break

    if errors:
        printReport(
            conn,
            host,
            "Test 6 failed for " + host + " - Found interfaces bouncing in recent log entries: \n",
            errors,
            commandsReport,
            deviceType,
        )
        return False

    else:

        printReport(
            conn,
            host,
            "Test 6 successfull for " + host + " - No recent log entries for any of its interfaces",
            "",
            "",
            deviceType,
        )
        return True


def checkEnvironment(conn, host, deviceType):
    errors = {}
    commandsReport = {}
    if deviceType == "arista_eos":
        i = 1
        commands = ["show system environment power", "show system environment temperature"]
        for command in commands:
            try:
                output = conn.enable(command)

            except:
                print(
                    "\nWARNING: The command '"
                    + command
                    + "' failed for the device "
                    + host
                    + ". Check the command syntax or run it manually on the device. Skipping this command ...\n"
                )
                continue

            if command.find("power") != -1 and bool(output[0]["result"]) == True:
                for PSU in output[0]["result"]["powerSupplies"].keys():
                    if not (
                        (output[0]["result"]["powerSupplies"][PSU]["state"] == "ok")
                        and (
                            val == "ok"
                            for val in set(
                                output[0]["result"]["powerSupplies"][PSU]["tempSensors"][sensors]["status"]
                                for sensors in output[0]["result"]["powerSupplies"][PSU]["tempSensors"].keys()
                            )
                        )
                        and (
                            val == "ok"
                            for val in set(
                                output[0]["result"]["powerSupplies"][PSU]["fans"][fans]["status"]
                                for fans in output[0]["result"]["powerSupplies"][PSU]["fans"].keys()
                            )
                        )
                    ):

                        errors[i] = {
                            "psuNumber": output[0]["result"]["powerSupplies"][PSU],
                            "psuState": output[0]["result"]["powerSupplies"][PSU]["state"],
                        }
                        errors[i].update(
                            dict(
                                (sensors, output[0]["result"]["powerSupplies"][PSU]["tempSensors"][sensors]["status"])
                                for sensors in output[0]["result"]["powerSupplies"][PSU]["tempSensors"].keys()
                            )
                        )
                        errors[i].update(
                            dict(
                                (fans, output[0]["result"]["powerSupplies"][PSU]["fans"][fans]["status"])
                                for fans in output[0]["result"]["powerSupplies"][PSU]["fans"].keys()
                            )
                        )
                        commandsReport[i] = {1: command}
                        i = i + 1

            if command.find("temperature") != -1 and bool(output[0]["result"]) == True:

                if not ((output[0]["result"]["systemStatus"] == "temperatureOk")):

                    for PSS in output[0]["result"].keys():

                        if PSS == "powerSupplySlots":

                            for PowerSupply in output[0]["result"][PSS]:

                                errors[i] = {
                                    "psuNumber": PowerSupply["relPos"],
                                }
                                errors[i].update(
                                    dict(
                                        (
                                            tempSensors["name"],
                                            "Description: "
                                            + str(tempSensors["description"])
                                            + " - Sensor status: "
                                            + str(tempSensors["hwStatus"])
                                            + " - Current temperature: "
                                            + str(tempSensors["currentTemperature"])
                                            + " - overheat Threshold: "
                                            + str(tempSensors["overheatThreshold"])
                                            + " - Alert state: "
                                            + str(tempSensors["inAlertState"]),
                                        )
                                        for tempSensors in PowerSupply["tempSensors"]
                                    )
                                )
                                commandsReport[i] = {1: command}
                                i = i + 1

                        elif PSS == "tempSensors":

                            for chassisTempSensors in output[0]["result"][PSS]:

                                errors[i] = {
                                    "Sensor Name": chassisTempSensors["name"],
                                    "Description": chassisTempSensors["description"],
                                    "Sensor status": chassisTempSensors["hwStatus"],
                                    "Current temperature": chassisTempSensors["currentTemperature"],
                                    "Overheat Threshold": chassisTempSensors["overheatThreshold"],
                                    "Alert state": chassisTempSensors["inAlertState"],
                                }
                                commandsReport[i] = {1: command}
                                i = i + 1

    if deviceType == "juniper_junos":

        command = "show chassis environment"
        try:
            Info = conn.rpc.get_environment_information()

        except:
            print(
                "\nWARNING: The command "
                + command
                + " failed for the device "
                + host
                + ". Check the command syntax or run it manually on the device. Skipping Test 6 ...\n"
            )
            return

        i = 1
        for item in Info.findall(".//environment-item"):
            if item.find(".//status").text.strip("\n") != "OK":
                if item.find(".//class").text.strip("\n") == "Fans":
                    errors[i] = {
                        "Sensor Name": (
                            item.find(".//name").text.strip("\n") if item.find(".//name") is not None else ""
                        ),
                        "Class": (item.find(".//class").text.strip("\n") if item.find(".//class") is not None else ""),
                        "Sensor status": (
                            item.find(".//status").text.strip("\n") if item.find(".//status") is not None else ""
                        ),
                        "Comment": (
                            item.find(".//comment").text.strip("\n") if item.find(".//comment") is not None else ""
                        ),
                    }

                else:

                    errors[i] = {
                        "Sensor Name": (
                            item.find(".//name").text.strip("\n") if item.find(".//name") is not None else ""
                        ),
                        "Sensor status": (
                            item.find(".//status").text.strip("\n") if item.find(".//status") is not None else ""
                        ),
                        "Comment": (
                            item.find(".//temperatute").text.strip("\n")
                            if item.find(".//temperatute") is not None
                            else ""
                        ),
                    }

                commandsReport[i] = {1: command}
                i = i + 1

        command = "show chassis alarms"
        try:
            Info = conn.rpc.get_alarm_information()

        except:
            print(
                "\nWARNING: The command "
                + command
                + " failed for the device "
                + host
                + ". Check the command syntax or run it manually on the device. Skipping Test 6 ...\n"
            )
            return

        for item in Info.findall(".//alarm-summary"):
            if item.find(".//active-alarm-count") is not None:
                for alarm in Info.findall(".//alarm-detail"):
                    if alarm.find(".//alarm-class").text == "Major":
                        errors[i] = {
                            "Chassis Alarm": (
                                alarm.find(".//alarm-description").text.strip("\n")
                                if alarm.find(".//alarm-description") is not None
                                else ""
                            ),
                            "Alarm time": (
                                alarm.find(".//alarm-time").text.strip("\n")
                                if alarm.find(".//alarm-time") is not None
                                else ""
                            ),
                        }
                        commandsReport[i] = {1: command}
                        i = i + 1

    if deviceType == "sr_linux":  # Nokia
        command = "/platform"
        try:
            platformInfo = getNokiaInformation(host, command)

        except:
            print(
                "\nWARNING: The command "
                + command
                + " failed for the device "
                + host
                + ". Check the command syntax or run it manually on the device. Skipping Test 3 ...\n"
            )
            return
        i = 1
        for platComponentsKey in platformInfo["notification"][0]["update"][0]["val"].keys():
            if platComponentsKey.find("fan-tray") != -1:
                for platComponents in platformInfo["notification"][0]["update"][0]["val"][platComponentsKey]:
                    if platComponents["oper-state"] != "up" and platComponents["oper-state"] != "empty":
                        errors[i] = {
                            "Fan-tray status": (
                                platComponents["oper-state"] if platComponents["oper-state"] is not None else "Unknown"
                            ),
                            "Fan-tray ID": (platComponents["id"] if platComponents["id"] is not None else "unknown"),
                            "Fan-tray type": (
                                platComponents["type"] if platComponents["type"] is not None else "Unknown"
                            ),
                        }

                    elif platComponents["oper-state"] == "empty":
                        errors[i] = platformInfo["notification"][0]["update"][0]["val"][platComponentsKey]

                    commandsReport[i] = {1: command}
                    i = i + 1
            elif platComponentsKey.find("linecard") != -1:
                for platComponents in platformInfo["notification"][0]["update"][0]["val"][platComponentsKey]:
                    if platComponents["oper-state"] != "up" and platComponents["oper-state"] != "empty":
                        errors[i] = {
                            "Linecard status": (
                                platComponents["oper-state"] if platComponents["oper-state"] is not None else "Unknown"
                            ),
                            "Linecard SN": (
                                platComponents["serial-number"]
                                if platComponents["serial-number"] is not None
                                else "unknown"
                            ),
                            "Linecard type": (
                                platComponents["type"] if platComponents["type"] is not None else "Unknown"
                            ),
                            "Linecard software version": (
                                platComponents["software version"]
                                if platComponents["software version"] is not None
                                else "Unknown"
                            ),
                            "Linecard slot": (
                                platComponents["slot"] if platComponents["slot"] is not None else "Unknown"
                            ),
                            "Linecard part-number": (
                                platComponents["part-number"]
                                if platComponents["part-number"] is not None
                                else "Unknown"
                            ),
                        }

                    elif platComponents["oper-state"] == "empty":
                        errors[i] = platformInfo["notification"][0]["update"][0]["val"][platComponentsKey]

                    commandsReport[i] = {1: command}
                    i = i + 1

            elif platComponentsKey.find("power-supply") != -1:
                for platComponents in platformInfo["notification"][0]["update"][0]["val"][platComponentsKey]:
                    if (
                        platComponents["oper-state"] != "up" or platComponents["temperature"]["alarm-status"]
                    ) and platComponents["oper-state"] != "empty":
                        errors[i] = {
                            "Power Supply status": (
                                platComponents["oper-state"] if platComponents["oper-state"] is not None else "Unknown"
                            ),
                            "Power Supply Temperature Alarm": (
                                platComponents["temperature"]["alarm-status"]
                                if platComponents["temperature"]["alarm-status"] is not None
                                else "Unknown"
                            ),
                            "Power Supply ID": (
                                platComponents["id"] if platComponents["id"] is not None else "Unknown"
                            ),
                            "Power Supply SN": (
                                platComponents["serial-number"]
                                if platComponents["serial-number"] is not None
                                else "unknown"
                            ),
                            "Power Supply type": (
                                platComponents["type"] if platComponents["type"] is not None else "Unknown"
                            ),
                            "Power Supply hot-swappable": (
                                platComponents["removable"] if platComponents["removable"] is not None else "Unknown"
                            ),
                            "Power Supply part-number": (
                                platComponents["part-number"]
                                if platComponents["part-number"] is not None
                                else "Unknown"
                            ),
                        }

                    elif platComponents["oper-state"] == "empty":
                        errors[i] = platformInfo["notification"][0]["update"][0]["val"][platComponentsKey]

                    commandsReport[i] = {1: command}
                    i = i + 1

            elif platComponentsKey.find("chassis") != -1:
                if platformInfo["notification"][0]["update"][0]["val"][platComponentsKey]["oper-state"] != "up":
                    errors[i] = {
                        "Chassis status": (
                            platformInfo["notification"][0]["update"][0]["val"][platComponentsKey]["oper-state"]
                            if platformInfo["notification"][0]["update"][0]["val"][platComponentsKey]["oper-state"]
                            is not None
                            else "Unknown"
                        ),
                        "Chassis Temperature Alarm": (
                            platComponents["temperature"]["alarm-status"]
                            if platComponents["temperature"]["alarm-status"] is not None
                            else "Unknown"
                        ),
                        "Chassis HW MAC Addr": (
                            platformInfo["notification"][0]["update"][0]["val"][platComponentsKey]["hw-mac-address"]
                            if platformInfo["notification"][0]["update"][0]["val"][platComponentsKey]["hw-mac-address"]
                            is not None
                            else "Unknown"
                        ),
                        "Chassis SN": (
                            platformInfo["notification"][0]["update"][0]["val"][platComponentsKey]["serial-number"]
                            if platformInfo["notification"][0]["update"][0]["val"][platComponentsKey]["serial-number"]
                            is not None
                            else "unknown"
                        ),
                        "Chassis type": (
                            platformInfo["notification"][0]["update"][0]["val"][platComponentsKey]["type"]
                            if platformInfo["notification"][0]["update"][0]["val"][platComponentsKey]["type"]
                            is not None
                            else "Unknown"
                        ),
                        "Chassis hot-swappable": (
                            platformInfo["notification"][0]["update"][0]["val"][platComponentsKey]["removable"]
                            if platformInfo["notification"][0]["update"][0]["val"][platComponentsKey]["removable"]
                            is not None
                            else "Unknown"
                        ),
                        "Chassis part-number": (
                            platformInfo["notification"][0]["update"][0]["val"][platComponentsKey]["part-number"]
                            if platformInfo["notification"][0]["update"][0]["val"][platComponentsKey]["part-number"]
                            is not None
                            else "Unknown"
                        ),
                    }
                    commandsReport[i] = {1: command}
                    i = i + 1

    if errors:
        printReport(
            conn,
            host,
            "Test 7 failed for " + host + " - High temperature warning or failed PSU found: \n",
            errors,
            commandsReport,
            deviceType,
        )
        return False
    else:
        printReport(
            conn, host, "Test 7 successfull for " + host + " - Temperature and power units are OK", "", "", deviceType
        )
        return True


def getBGPNeigAdapterJuniper(BGPNeig, command):
    # Fuction used to be able to re-use the code for both Juniper and Arista devices in the function checkBGPNeighbors
    return_dict = {}
    i = 1
    for item in BGPNeig.findall(".//bgp-peer"):

        if item.find(".//description") is not None:
            if item.find(".//description").text.find("CUST") == -1:  # Skip customers' BGP sessions
                return_dict[i] = {
                    "peerAddr": item.find(".//peer-address").text.strip("\n"),
                    "status": item.find(".//peer-state").text.strip("\n"),
                    "description": item.find(".//description").text.strip("\n"),
                    "ASN": item.find(".//peer-as").text.strip("\n"),
                    "command": "show bgp neighbor " + str(item.find(".//peer-address").text),
                }

        else:
            return_dict[i] = {
                "peerAddr": item.find(".//peer-address").text.strip("\n"),
                "status": item.find(".//peer-state").text.strip("\n"),
                "description": "",
                "ASN": item.find(".//peer-as").text.strip("\n"),
                "command": "show bgp neighbor " + str(item.find(".//peer-address").text),
            }

        i = i + 1
    return return_dict


def getBGPNeigAdapterArista(BGPNeig, command):
    # Fuction used to be able to re-use the code for both Juniper and Arista devices in the function checkBGPNeighbors
    return_dict = {}
    i = 1
    for item in BGPNeig["vrfs"]["default"]["peers"].keys():

        if "description" in BGPNeig["vrfs"]["default"]["peers"][item]:
            if (
                BGPNeig["vrfs"]["default"]["peers"][item]["description"].find("CUST") == -1
            ):  # Skip customers' BGP sessions
                return_dict[i] = {
                    "peerAddr": item,
                    "status": BGPNeig["vrfs"]["default"]["peers"][item]["peerState"],
                    "description": BGPNeig["vrfs"]["default"]["peers"][item]["description"],
                    "ASN": BGPNeig["vrfs"]["default"]["peers"][item]["asn"],
                    "command": command,
                }

        else:
            return_dict[i] = {
                "peerAddr": item,
                "status": BGPNeig["vrfs"]["default"]["peers"][item]["peerState"],
                "description": "",
                "ASN": BGPNeig["vrfs"]["default"]["peers"][item]["asn"],
                "command": command,
            }

        i = i + 1
    return return_dict


def getBGPNeigAdapterNokia(BGPNeig, command):
    # Fuction used to be able to re-use the code for Juniper, Nokia and Arista devices in the function checkBGPNeighbors
    return_dict = {}
    i = 1
    for item in BGPNeig["notification"][0]["update"][0]["val"]["neighbor"]:
        if "description" in item.keys():
            if item["description"].find("CUST") == -1:  # Skip customers' BGP sessions
                return_dict[i] = {
                    "peerAddr": item["peer-address"],
                    "status": item["session-state"],
                    "description": item["description"] if item["description"] is not None else "",
                    "ASN": item["peer-as"],
                    "command": str(command.split("[peer")[0]) + "[peer-address=" + str(item["peer-address"]) + "]",
                }

        else:
            return_dict[i] = {
                "peerAddr": item["peer-address"],
                "status": item["session-state"],
                "description": "",
                "ASN": item["peer-as"],
                "command": str(command.split("[peer")[0]) + "[peer-address=" + str(item["peer-address"]) + "]",
            }

        i = i + 1
    return return_dict


def checkBGPNeighbors(conn, host, deviceType):
    # This fucntion checks that the bgp neighbors for both the ipv4 and evpn address families are UP. If they all are UP it returns True and if not,
    # it returns False and prints the errors found in the report
    errors = {}
    commandsReport = {}
    BGPPeerInfo = {}
    i = 1
    if deviceType == "arista_eos":  # Arista
        commands = ["show bgp ipv4 unicast summary", "show bgp ipv6 unicast summary", "show bgp evpn summary"]
        for command in commands:
            try:
                output = conn.enable(command)

            except:
                print(
                    "\nWARNING: The command '"
                    + command
                    + "' failed for the device "
                    + host
                    + ". Check the command syntax or run it manually on the device. Skipping this command\n"
                )
                continue

            if bool(output[0]["result"]) == True:
                BGPPeerInfo.update(getBGPNeigAdapterArista(output[0]["result"], command))

    if deviceType == "juniper_junos":  # Juniper

        command = "show bgp summary"
        try:
            BGPInfo = conn.rpc.get_bgp_summary_information()

        except:
            print(
                "\nWARNING: The command "
                + command
                + " failed for the device "
                + host
                + ". Check the command syntax or run it manually on the device. Skipping Test 6 ...\n"
            )
            return

        BGPPeerInfo = getBGPNeigAdapterJuniper(BGPInfo, command)

    if deviceType == "sr_linux":  # Nokia
        command = "network-instance[name=default]/protocols/bgp/neighbor[peer-address=*]"
        try:
            BGPInfo = getNokiaInformation(host, command)

        except:
            print(
                "\nWARNING: The command "
                + command
                + " failed for the device "
                + host
                + ". Check the command syntax or run it manually on the device. Skipping Test 6 ...\n"
            )
            return

        BGPPeerInfo = getBGPNeigAdapterNokia(BGPInfo, command)

    for peerNbr in BGPPeerInfo.keys():

        if not (BGPPeerInfo[peerNbr]["status"].lower() == "established"):

            errors[i] = {
                "hostname": host,
                "bgpNeighborDown": BGPPeerInfo[peerNbr]["peerAddr"],
                "bgpPeerDescription": BGPPeerInfo[peerNbr]["description"],
                "bgpPeerASN": BGPPeerInfo[peerNbr]["ASN"],
            }
            commandsReport[i] = {1: BGPPeerInfo[peerNbr]["command"]}
            i += 1

    if errors:
        printReport(
            conn,
            host,
            "Test 8 failed for " + host + " - Found BGP neighbors in down state: \n",
            errors,
            commandsReport,
            deviceType,
        )
        return False

    else:
        printReport(
            conn,
            host,
            "Test 8 successfull for " + host + " - All the IPv4, IPv6 and EVPN bgp neighbors are UP",
            "",
            "",
            deviceType,
        )
        return True


def getInterfaceMTU(conn, host, interface, deviceType):
    # Fuction that returns the mtu value that is used in the test CheckInterfaceErrors for the ping commands.

    if deviceType == "arista_eos":

        command = "show interfaces " + str(interface)

        try:
            output = conn.enable(command)

        except:
            print(
                "\nWARNING: The command '"
                + command
                + "' failed for the device "
                + host
                + ". Check the command syntax or run it manually on the device. Skipping the test of the following interface: "
                + interface
                + "\n"
            )
            return int(1500) - 60

        if interface in output[0]["result"]["interfaces"].keys():
            return int(output[0]["result"]["interfaces"][interface]["mtu"]) - 60

        else:
            return int(1500) - 60

    if deviceType == "juniper_junos":
        command = "show interfaces extensive " + str(interface)
        try:
            intInfo = conn.rpc.get_interface_information(extensive=True, interface_name=str(interface).split(".")[0])

        except:
            print(
                "\nWARNING: The command '"
                + command
                + "' failed for the device "
                + host
                + ". Check the command syntax or run it manually on the device. Skipping this command...\n"
            )
            return int(1500) - 60

        if intInfo.find(".//description") is not None:
            if "ECX" in intInfo.find(".//description").text:  # Max MTU supported by fabric is 9100 bytes
                return int(9100) - 60

            else:
                return int(intInfo.find(".//mtu").text.strip("\n")) - 60
        else:
            return int(1350)

    if deviceType == "sr_linux":
        command = "/interface[name=" + str(interface) + "]"
        try:
            intInfo = getNokiaInformation(host, command)

        except:
            print(
                "\nWARNING: The command '"
                + command
                + "' failed for the device "
                + host
                + ". Check the command syntax or run it manually on the device. Skipping this command...\n"
            )
            return int(1500) - 60

        if intInfo["notification"][0]["update"][0]["val"]["mtu"]:
            return int(intInfo["notification"][0]["update"][0]["val"]["mtu"]) - 60

        else:
            return int(1500) - 60


def checkInterfaceErrors(conn, host, deviceType):
    # Function that generates some traffic with ping commands and checks the interface errors of all the device interfaces. If no errors found it returns True and if not,
    # it returns False and prints the errors found in the report
    interfacesNB = getDevInterfacesNB(host)
    errors = {}
    return_dict = {}
    commandsReport = {}
    i = 1

    if deviceType == "arista_eos":  # Arista

        for n in interfacesNB.keys():

            command = "clear counters " + str(interfacesNB[n]["localInterface"])

            try:
                output = conn.enable(command)

            except:
                print(
                    "\nWARNING: The command '"
                    + command
                    + "' failed for the device "
                    + host
                    + ". Check the command syntax or run it manually on the device. Skipping the test of the following interface: "
                    + interfacesNB[n]["localInterface"]
                    + "\n"
                )
                errors[i] = {
                    "Hostname": host,
                    "local Interface": interfacesNB[n]["localInterface"],
                    "Command": command,
                    "error": "The command '"
                    + command
                    + "' failed for the device "
                    + host
                    + ". Check the command syntax or run it manually on the device. Skipping the test of the following interface: "
                    + interfacesNB[n]["localInterface"]
                    + "\n",
                }
                commandsReport[i] = {
                    1: command,
                    2: "show ip interface " + interfacesNB[n]["localInterface"] + " brief",
                    3: "show interfaces " + interfacesNB[n]["localInterface"] + " status",
                    4: "show interfaces " + interfacesNB[n]["localInterface"] + " counters errors",
                }
                i = i + 1
                continue

            InterfaceIPAddr = getInterfaceIPaddresses(host, interfacesNB[n]["localInterfaceID"])
            if bool(InterfaceIPAddr) == True:
                ipv4Addr = ipaddress.IPv4Address(InterfaceIPAddr[1]["Address"].split("/")[0])
                ipv4octets = InterfaceIPAddr[1]["Address"].split("/")[0].split(".")
                if (int(ipv4octets[3]) % 2) == 0:
                    nextHopIPaddress = ipv4Addr + 1
                else:
                    nextHopIPaddress = ipv4Addr - 1

                if interfacesNB[n]["localInterface"] == "Management1":

                    command = (
                        "ping vrf MGMT "
                        + str(nextHopIPaddress)
                        + " source "
                        + str(ipv4Addr)
                        + " size "
                        + str(getInterfaceMTU(conn, host, interfacesNB[n]["localInterface"], deviceType))
                        + " repeat 1"
                    )

                else:

                    command = (
                        "ping "
                        + str(nextHopIPaddress)
                        + " source "
                        + str(ipv4Addr)
                        + " size "
                        + str(getInterfaceMTU(conn, host, interfacesNB[n]["localInterface"], deviceType))
                        + " repeat 1"
                    )

                # check if the ping works with one probe (otherwise, it delays the whole script a lot)
                try:
                    output = conn.enable(command)

                except:
                    print(
                        "\nWARNING: The command '"
                        + command
                        + "' failed for the device "
                        + host
                        + ". Check the command syntax or run it manually on the device. Skipping the test of the following interface: "
                        + interfacesNB[n]["localInterface"]
                        + "\n"
                    )
                    errors[i] = {
                        "Hostname": host,
                        "local Interface": interfacesNB[n]["localInterface"],
                        "Command": command,
                        "error": "The command '"
                        + command
                        + "' failed for the device "
                        + host
                        + ". Check the command syntax or run it manually on the device. Skipping the test of the following interface: "
                        + interfacesNB[n]["localInterface"]
                        + "\n",
                    }
                    commandsReport[i] = {
                        1: command,
                        2: "show ip interface " + interfacesNB[n]["localInterface"] + " brief",
                        3: "show interfaces " + interfacesNB[n]["localInterface"] + " status",
                        4: "show interfaces " + interfacesNB[n]["localInterface"] + " counters errors",
                    }
                    i += 1
                    continue

                if bool(output[0]["result"]) == True:
                    pings = str(output[0]["result"]).split("\\n")
                    errorFound = False
                    for line in pings:  # checking errors in the response of the probe
                        if "bind: Cannot assign requested address" in line:
                            errors[i] = {
                                "Hostname": host,
                                "local Interface": interfacesNB[n]["localInterface"],
                                "Command": command,
                                "error": line,
                            }
                            commandsReport[i] = {
                                1: command,
                                2: "show ip interface " + interfacesNB[n]["localInterface"] + " brief",
                                3: "show interfaces " + interfacesNB[n]["localInterface"] + " status",
                                4: "show interfaces " + interfacesNB[n]["localInterface"] + " counters errors",
                            }
                            i += 1
                            errorFound = True
                            continue

                        elif (
                            ("1 packets transmitted, 1 packets received, 0% packet loss" not in line)
                            and ("0 packets received" in line)
                        ) and line:
                            errors[i] = {
                                "Hostname": host,
                                "local Interface": interfacesNB[n]["localInterface"],
                                "error": "The ping command failed - 100% packet loss when pinging the other end",
                            }
                            commandsReport[i] = {
                                1: command,
                                2: "show ip interface " + interfacesNB[n]["localInterface"] + " brief",
                                3: "show interfaces " + interfacesNB[n]["localInterface"] + " status",
                                4: "show interfaces " + interfacesNB[n]["localInterface"] + " counters errors",
                            }
                            i += 1
                            errorFound = True
                            continue

                    if not errorFound:  # ping worked. Sending a large amount of pings

                        if interfacesNB[n]["localInterface"] == "Management1":

                            command = (
                                "ping vrf MGMT "
                                + str(nextHopIPaddress)
                                + " source "
                                + str(ipv4Addr)
                                + " size "
                                + str(getInterfaceMTU(conn, host, interfacesNB[n]["localInterface"], deviceType))
                                + " repeat 1000"
                            )

                        else:

                            command = (
                                "ping "
                                + str(nextHopIPaddress)
                                + " source "
                                + str(ipv4Addr)
                                + " size "
                                + str(getInterfaceMTU(conn, host, interfacesNB[n]["localInterface"], deviceType))
                                + " repeat 1000"
                            )

                        try:
                            output = conn.enable(command)

                        except:
                            print(
                                "\nWARNING: The command '"
                                + command
                                + "' failed for the device "
                                + host
                                + ". Check the command syntax or run it manually on the device. Skipping the test of the following interface: "
                                + interfacesNB[n]["localInterface"]
                                + "\n"
                            )
                            errors[i] = {
                                "Hostname": host,
                                "local Interface": interfacesNB[n]["localInterface"],
                                "Command": command,
                                "error": "The command '"
                                + command
                                + "' failed for the device "
                                + host
                                + ". Check the command syntax or run it manually on the device. Skipping the test of the following interface: "
                                + interfacesNB[n]["localInterface"]
                                + "\n",
                            }
                            commandsReport[i] = {
                                1: command,
                                2: "show ip interface " + interfacesNB[n]["localInterface"] + " brief",
                                3: "show interfaces " + interfacesNB[n]["localInterface"] + " status",
                                4: "show interfaces " + interfacesNB[n]["localInterface"] + " counters errors",
                            }
                            i += 1
                            continue
                # Checking interface errors if pings are successful
                command = "show interfaces " + str(interfacesNB[n]["localInterface"]) + " counters errors"
                try:
                    output = conn.enable(command)

                except:
                    print(
                        "\nWARNING: The command '"
                        + command
                        + "' failed for the device "
                        + host
                        + ". Check the command syntax or run it manually on the device. Skipping this command...\n"
                    )
                    continue

                if (
                    bool(output[0]["result"]) == True
                    and interfacesNB[n]["localInterface"] in output[0]["result"]["interfaceErrorCounters"].keys()
                ):

                    for err in output[0]["result"]["interfaceErrorCounters"][interfacesNB[n]["localInterface"]]:

                        if output[0]["result"]["interfaceErrorCounters"][interfacesNB[n]["localInterface"]][err] > 0:
                            errors[i] = {
                                "Hostname": host,
                                "local Interface": interfacesNB[n]["localInterface"],
                                "error": str(err)
                                + ": "
                                + str(
                                    output[0]["result"]["interfaceErrorCounters"][interfacesNB[n]["localInterface"]][
                                        err
                                    ]
                                ),
                            }
                            commandsReport[i] = {
                                1: command,
                                2: "show ip interface " + interfacesNB[n]["localInterface"] + " brief",
                                3: "show interfaces " + interfacesNB[n]["localInterface"] + " status",
                                4: "show interfaces " + interfacesNB[n]["localInterface"] + " counters errors",
                                5: "show interfaces " + interfacesNB[n]["localInterface"],
                            }
                            i += 1

    if deviceType == "juniper_junos":  # Juniper

        for n in interfacesNB.keys():

            command = "clear interfaces statistics " + str(interfacesNB[n]["localInterface"])

            try:

                ClearIntCountersInfo = getDeviceOutput(host, command)

            except:
                print(
                    "\nWARNING: The command '"
                    + command
                    + "' failed for the device "
                    + host
                    + ". Check the command syntax or run it manually on the device. Skipping the test of the following interface: "
                    + interfacesNB[n]["localInterface"]
                    + "\n"
                )
                continue

            InterfaceIPAddr = getInterfaceIPaddresses(host, interfacesNB[n]["localInterfaceID"])

            if bool(InterfaceIPAddr) == True and interfacesNB[n]["localInterface"] != "fxp0":
                ipv4Addr = ipaddress.IPv4Address(InterfaceIPAddr[1]["Address"].split("/")[0])
                ipv4octets = InterfaceIPAddr[1]["Address"].split("/")[0].split(".")
                if (int(ipv4octets[3]) % 2) == 0:
                    nextHopIPaddress = ipv4Addr + 1
                else:
                    nextHopIPaddress = ipv4Addr - 1

                if (
                    interfacesNB[n]["localInterface"] == "xe-0/1/11:0"
                    or interfacesNB[n]["localInterface"] == "xe-1/1/11:0"
                ):

                    command = (
                        "ping routing-instance PACKET-INTERNAL "
                        + str(nextHopIPaddress)
                        + " source "
                        + str(ipv4Addr)
                        + " size "
                        + str(getInterfaceMTU(conn, host, interfacesNB[n]["localInterface"], deviceType))
                        + " rapid count 1 interval 2"
                    )

                else:
                    command = (
                        "ping "
                        + str(nextHopIPaddress)
                        + " source "
                        + str(ipv4Addr)
                        + " size "
                        + str(getInterfaceMTU(conn, host, interfacesNB[n]["localInterface"], deviceType))
                        + " rapid count 1 interval 2"
                    )

                # check if the ping works with one probe (otherwise it delays the whole script a lot)
                try:

                    pingInfo = getDeviceOutput(host, command)

                except:
                    print(
                        "\nWARNING: The command '"
                        + command
                        + "' failed for the device "
                        + host
                        + ". Check the command syntax or run it manually on the device. Skipping the test of the following interface: "
                        + interfacesNB[n]["localInterface"]
                        + "\n"
                    )
                    errors[i] = {
                        "Hostname": host,
                        "local Interface": interfacesNB[n]["localInterface"],
                        "Command": command,
                        "error": "The command '"
                        + command
                        + "' failed for the device "
                        + host
                        + ". Check the command syntax or run it manually on the device. Skipping the test of the following interface: "
                        + interfacesNB[n]["localInterface"]
                        + "\n",
                    }
                    commandsReport[i] = {
                        1: command,
                        2: "show interfaces " + interfacesNB[n]["localInterface"] + " descriptions",
                        3: "show interfaces extensive " + interfacesNB[n]["localInterface"],
                    }
                    i += 1
                    continue

                lines = str(pingInfo).splitlines()
                errorFound = False
                for line in lines:  # checkf for errors
                    if "bind: Can't assign requested address" in line and line:
                        errors[i] = {
                            "Hostname": host,
                            "local Interface": interfacesNB[n]["localInterface"],
                            "error": "The ping command failed - bind: Cannot assign requested address",
                        }
                        commandsReport[i] = {
                            1: command,
                            2: "show interfaces " + interfacesNB[n]["localInterface"] + " descriptions",
                            3: "show interfaces extensive " + interfacesNB[n]["localInterface"],
                        }
                        i += 1
                        errorFound = True
                        continue

                    elif (
                        ("1 packets transmitted, 1 packets received, 0% packet loss" not in line)
                        and ("0 packets received" in line)
                    ) and line:
                        errors[i] = {
                            "Hostname": host,
                            "local Interface": interfacesNB[n]["localInterface"],
                            "error": "The ping command failed - 100% packet loss when pinging the other end",
                        }
                        commandsReport[i] = {
                            1: command,
                            2: "show interfaces " + interfacesNB[n]["localInterface"] + " descriptions",
                            3: "show interfaces extensive " + interfacesNB[n]["localInterface"],
                        }
                        i += 1
                        errorFound = True
                        continue

                if not errorFound:  # ping successful. Send a large amount of pings
                    if (
                        interfacesNB[n]["localInterface"] == "xe-0/1/11:0"
                        or interfacesNB[n]["localInterface"] == "xe-1/1/11:0"
                    ):
                        command = (
                            "ping routing-instance PACKET-INTERNAL "
                            + str(nextHopIPaddress)
                            + " source "
                            + str(ipv4Addr)
                            + " size "
                            + str(getInterfaceMTU(conn, host, interfacesNB[n]["localInterface"], deviceType))
                            + " rapid count 1000"
                        )

                    else:

                        command = (
                            "ping "
                            + str(nextHopIPaddress)
                            + " source "
                            + str(ipv4Addr)
                            + " size "
                            + str(getInterfaceMTU(conn, host, interfacesNB[n]["localInterface"], deviceType))
                            + " rapid count 1000"
                        )

                    try:

                        pingInfo = getDeviceOutput(host, command)

                    except:
                        print(
                            "\nWARNING: The command '"
                            + command
                            + "' failed for the device "
                            + host
                            + ". Check the command syntax or run it manually on the device. Skipping the test of the following interface: "
                            + interfacesNB[n]["localInterface"]
                            + "\n"
                        )
                        errors[i] = {
                            "Hostname": host,
                            "local Interface": interfacesNB[n]["localInterface"],
                            "Command": command,
                            "error": "The command '"
                            + command
                            + "' failed for the device "
                            + host
                            + ". Check the command syntax or run it manually on the device. Skipping the test of the following interface: "
                            + interfacesNB[n]["localInterface"]
                            + "\n",
                        }
                        commandsReport[i] = {
                            1: command,
                            2: "show interfaces " + interfacesNB[n]["localInterface"] + " terse",
                            3: "show interfaces " + interfacesNB[n]["localInterface"] + " descriptions",
                            4: "show interfaces extensive " + interfacesNB[n]["localInterface"],
                        }
                        i += 1
                        continue

                # Checking interface errors if pings are successful
                command = "show interfaces extensive " + str(interfacesNB[n]["localInterface"]).split(".")[0]
                try:
                    intInfo = conn.rpc.get_interface_information(
                        extensive=True, interface_name=str(interfacesNB[n]["localInterface"]).split(".")[0]
                    )

                except:
                    print(
                        "\nWARNING: The command '"
                        + command
                        + "' failed for the device "
                        + host
                        + ". Check the command syntax or run it manually on the device. Skipping this command...\n"
                    )
                    continue

                if (
                    len(intInfo) > 0 and interfacesNB[n]["localInterface"].find("em") == -1
                ):  # For interfaces different from the em<x> because some counters are not available and break the script when testing the mrrs (Juniper MX204)
                    return_dict[str(interfacesNB[n]["localInterface"])] = {}
                    return_dict[str(interfacesNB[n]["localInterface"])] = {
                        0: {
                            "carrier-transitions": intInfo.find(".//carrier-transitions").text.strip("\n"),
                            "output-errors": intInfo.find(".//output-errors").text.strip("\n"),
                            "output-collisions": intInfo.find(".//output-collisions").text.strip("\n"),
                            "output-drops": intInfo.find(".//output-drops").text.strip("\n"),
                            "aged-packets": intInfo.find(".//aged-packets").text.strip("\n"),
                            "mtu-errors": intInfo.find(".//mtu-errors").text.strip("\n"),
                            "hs-link-crc-errors": intInfo.find(".//hs-link-crc-errors").text.strip("\n"),
                            "output-fifo-errors": intInfo.find(".//output-fifo-errors").text.strip("\n"),
                            "output-resource-errors": intInfo.find(".//output-resource-errors").text.strip("\n"),
                            "input-errors": intInfo.find(".//input-errors").text.strip("\n"),
                            "framing-errors": intInfo.find(".//framing-errors").text.strip("\n"),
                            "input-runts": intInfo.find(".//input-runts").text.strip("\n"),
                            "input-discards": intInfo.find(".//input-discards").text.strip("\n"),
                            "input-l3-incompletes": intInfo.find(".//input-l3-incompletes").text.strip("\n"),
                            "input-l2-channel-errors": intInfo.find(".//input-l2-channel-errors").text.strip("\n"),
                            "input-fifo-errors": intInfo.find(".//input-fifo-errors").text.strip("\n"),
                            "input-resource-errors": intInfo.find(".//input-resource-errors").text.strip("\n"),
                        }
                    }

                elif len(intInfo) and interfacesNB[n]["localInterface"].find("em") != -1:  # For em<x> interfaces
                    return_dict[str(interfacesNB[n]["localInterface"])] = {}
                    return_dict[str(interfacesNB[n]["localInterface"])] = {
                        0: {
                            "carrier-transitions": intInfo.find(".//carrier-transitions").text.strip("\n"),
                            "output-errors": intInfo.find(".//output-errors").text.strip("\n"),
                            "output-drops": intInfo.find(".//output-drops").text.strip("\n"),
                            "mtu-errors": intInfo.find(".//mtu-errors").text.strip("\n"),
                            "output-resource-errors": intInfo.find(".//output-resource-errors").text.strip("\n"),
                            "input-errors": intInfo.find(".//input-errors").text.strip("\n"),
                            "framing-errors": intInfo.find(".//framing-errors").text.strip("\n"),
                            "input-drops": intInfo.find(".//input-drops").text.strip("\n"),
                            "input-discards": intInfo.find(".//input-discards").text.strip("\n"),
                            "input-runts": intInfo.find(".//input-runts").text.strip("\n"),
                            "input-giants": intInfo.find(".//input-giants").text.strip("\n"),
                            "input-resource-errors": intInfo.find(".//input-resource-errors").text.strip("\n"),
                        }
                    }

                    for err in return_dict[str(interfacesNB[n]["localInterface"])][0]:

                        if int(return_dict[str(interfacesNB[n]["localInterface"])][0][err]) > 0:
                            errors[i] = {
                                "Hostname": host,
                                "local Interface": interfacesNB[n]["localInterface"],
                                "error": str(err)
                                + ": "
                                + str(return_dict[str(interfacesNB[n]["localInterface"])][0][err]),
                            }
                            commandsReport[i] = {1: command}
                            i += 1

    if deviceType == "sr_linux":  # Nokia
        nokiaIntErrors = [
            "carrier-transitions",
            "in-error-packets",
            "in-fcs-error-packets",
            "out-error-packets",
        ]  # Nokia error names to look for errors
        for n in interfacesNB.keys():

            command = "tools interface " + str(interfacesNB[n]["localInterface"]) + " statistics clear"

            try:
                output = getDeviceOutput(host, command)

            except:
                print(
                    "\nWARNING: The command '"
                    + command
                    + "' failed for the device "
                    + host
                    + ". Check the command syntax or run it manually on the device. Skipping the test of the following interface: "
                    + interfacesNB[n]["localInterface"]
                    + "\n"
                )
                continue

            InterfaceIPAddr = getInterfaceIPaddresses(host, interfacesNB[n]["localInterfaceID"])
            if bool(InterfaceIPAddr) == True:
                ipv4Addr = ipaddress.IPv4Address(InterfaceIPAddr[1]["Address"].split("/")[0])
                ipv4octets = InterfaceIPAddr[1]["Address"].split("/")[0].split(".")
                if (int(ipv4octets[3]) % 2) == 0:
                    nextHopIPaddress = ipv4Addr + 1
                else:
                    nextHopIPaddress = ipv4Addr - 1

                if interfacesNB[n]["localInterface"] == "mgmt0":

                    command = (
                        "ping network-instance mgmt -I "
                        + str(ipv4Addr)
                        + " "
                        + str(nextHopIPaddress)
                        + " -s "
                        + str(getInterfaceMTU(conn, host, interfacesNB[n]["localInterface"], deviceType))
                        + " -c 1"
                    )

                else:

                    command = (
                        "ping network-instance default -I "
                        + str(ipv4Addr)
                        + " "
                        + str(nextHopIPaddress)
                        + " -s "
                        + str(getInterfaceMTU(conn, host, interfacesNB[n]["localInterface"], deviceType))
                        + " -c 1"
                    )

                try:
                    output = getDeviceOutput(host, command)

                except:
                    print(
                        "\nWARNING: The command '"
                        + command
                        + "' failed for the device "
                        + host
                        + ". Check the command syntax or run it manually on the device. Skipping the test of the following interface: "
                        + interfacesNB[n]["localInterface"]
                        + "\n"
                    )
                    errors[i] = {
                        "Hostname": host,
                        "local Interface": interfacesNB[n]["localInterface"],
                        "Command": command,
                        "error": "The command '"
                        + command
                        + "' failed for the device "
                        + host
                        + ". Check the command syntax or run it manually on the device. Skipping the test of the following interface: "
                        + interfacesNB[n]["localInterface"]
                        + "\n",
                    }
                    commandsReport[i] = {
                        1: command,
                        2: "show interface " + interfacesNB[n]["localInterface"] + " detail",
                    }
                    i += 1
                    continue

                if output:
                    lines = str(pingInfo).splitlines()
                    errorFound = False
                    for line in lines:  # check for errors
                        if "bind: Can't assign requested address" in line and line:
                            errors[i] = {
                                "Hostname": host,
                                "local Interface": interfacesNB[n]["localInterface"],
                                "error": "The ping command failed - bind: Cannot assign requested address",
                            }
                            commandsReport[i] = {
                                1: command,
                                2: "show interface " + interfacesNB[n]["localInterface"] + " detail",
                            }
                            i += 1
                            errorFound = True
                            continue

                        elif (
                            ("1 packets transmitted, 1 packets received, 0% packet loss" not in line)
                            and ("0 packets received" in line)
                        ) and line:
                            errors[i] = {
                                "Hostname": host,
                                "local Interface": interfacesNB[n]["localInterface"],
                                "error": "The ping command failed - 100% packet loss when pinging the other end",
                            }
                            commandsReport[i] = {
                                1: command,
                                2: "show interface " + interfacesNB[n]["localInterface"] + " detail",
                            }
                            i += 1
                            errorFound = True
                            continue

                    if not errorFound:  # ping successful. Send a large amount of pings
                        if (
                            interfacesNB[n]["localInterface"] == "xe-0/1/11:0"
                            or interfacesNB[n]["localInterface"] == "xe-1/1/11:0"
                        ):
                            command = (
                                "ping routing-instance PACKET-INTERNAL "
                                + str(nextHopIPaddress)
                                + " source "
                                + str(ipv4Addr)
                                + " size "
                                + str(getInterfaceMTU(conn, host, interfacesNB[n]["localInterface"], deviceType))
                                + " rapid count 1000"
                            )

                        else:

                            command = (
                                "ping "
                                + str(nextHopIPaddress)
                                + " source "
                                + str(ipv4Addr)
                                + " size "
                                + str(getInterfaceMTU(conn, host, interfacesNB[n]["localInterface"], deviceType))
                                + " rapid count 1000"
                            )

                        try:

                            pingInfo = getDeviceOutput(host, command)

                        except:
                            print(
                                "\nWARNING: The command '"
                                + command
                                + "' failed for the device "
                                + host
                                + ". Check the command syntax or run it manually on the device. Skipping the test of the following interface: "
                                + interfacesNB[n]["localInterface"]
                                + "\n"
                            )
                            errors[i] = {
                                "Hostname": host,
                                "local Interface": interfacesNB[n]["localInterface"],
                                "Command": command,
                                "error": "The command '"
                                + command
                                + "' failed for the device "
                                + host
                                + ". Check the command syntax or run it manually on the device. Skipping the test of the following interface: "
                                + interfacesNB[n]["localInterface"]
                                + "\n",
                            }
                            commandsReport[i] = {
                                1: command,
                                2: "show interface " + interfacesNB[n]["localInterface"] + " detail",
                            }
                            i = i + 1
                            continue

                # Checking interface errors if pings are successful
                command = "interface[name=" + str(interfacesNB[n]["localInterface"]) + "]"
                try:
                    output = getNokiaInformation(host, command)

                except:
                    print(
                        "\nWARNING: The command '"
                        + command
                        + "' failed for the device "
                        + host
                        + ". Check the command syntax or run it manually on the device. Skipping this command...\n"
                    )
                    continue

                if bool(output["notification"][0]["update"][0]["val"]["statistics"]) == True:
                    for err in output["notification"][0]["update"][0]["val"]["statistics"]:
                        if err != "last-clear":
                            if (
                                int(output["notification"][0]["update"][0]["val"]["statistics"][err]) > 0
                                and err in nokiaIntErrors
                            ):
                                errors[i] = {
                                    "Hostname": host,
                                    "local Interface": interfacesNB[n]["localInterface"],
                                    "error": str(err)
                                    + ": "
                                    + str(output["notification"][0]["update"][0]["val"]["statistics"][err]),
                                }
                                commandsReport[i] = {
                                    1: command,
                                    2: "show interface " + interfacesNB[n]["localInterface"] + " detail",
                                }
                                i += 1

    if errors:
        printReport(
            conn,
            host,
            "Test 9 failed for " + host + " - Interface errors found or ping command failed: \n",
            errors,
            commandsReport,
            deviceType,
        )
        return False

    else:

        printReport(
            conn,
            host,
            "Test 9 successfull for " + host + " - Pings worked successfully and no interface errors found",
            "",
            "",
            deviceType,
        )
        return True


def checkMswDhcpRelay(conn, host, deviceType):
    # Function that checks if dhcp forwarding is correctly configured
    commandsReport = {}
    errors = {}
    i = 1
    if deviceType == "arista_eos" and host.startswith("msw"):
        command = "show ip helper-address"
        try:
            output = conn.enable(command)
        except:
            print(
                f"\nWARNING: The command {command} failed for the device {host}. Check the command syntax or run it manually on the device. Skipping this command ...\n"
            )

        if len(output[0]["result"]["configuredInterfaces"]["Vlan10"]["helpers"]) > 1:
            errors[i] = f"Test failed for {host} - There is more than one server set as the helper address on vlan 10"
            commandsReport[i] = {1: command}
            i = i + 1
        else:
            for server in output[0]["result"]["configuredInterfaces"]["Vlan10"]["helpers"]:
                if "TINKERBELL" not in server["serverHostname"]:
                    errors[i] = (
                        f"Test failed for {host} - TINKERBELL is not set as the helper address on vlan 10. It is set as {server['serverHostname']}"
                    )
                    commandsReport[i] = {1: command}

        if errors:
            printReport(
                conn,
                host,
                f"Test 17 failed for {host} - The dhcp helper on VLAN10 is configured incorrectly \n",
                errors,
                commandsReport,
                deviceType,
            )
            return False
        else:
            printReport(
                conn,
                host,
                f"Test 17 successfull both {host} -The dhcp helper on VLAN10 is configured correctly",
                "",
                "",
                deviceType,
            )
            return True

    elif deviceType == "juniper_junos" and host.startswith("msw"):
        junos_filter = "<configuration><forwarding-options/></configuration>"  # There is no direct command like "show forwding-options so we need to do show config and filter..."
        try:
            data = conn.rpc.get_config(filter_xml=junos_filter)
        except:
            print(
                f"\nWARNING: The command {command} failed for the device {host}. Check the command syntax or run it manually on the device. Skipping this command ...\n"
            )
        for group in data.xpath(".//group"):
            group_name = group.xpath(".//name")
            active_server_group = group.xpath(".//active-server-group/active-server-group")
            if group_name and group_name[0].text == "TINKERBELL":
                if not active_server_group or active_server_group[0].text != "TINKERBELL":
                    errors[i] = f"Test failed for {host} - TINKERBELL is not set as the dhcp relay"
                    commandsReport[i] = {1: command}
        if errors:
            printReport(
                conn,
                host,
                f"Test 17 failed for {host} - The dhcp helper is configured incorrectly \n",
                errors,
                commandsReport,
                deviceType,
            )
            return False
        else:
            printReport(
                conn,
                host,
                f"Test 17 successfull both {host} -The dhcp helper is configured correctly",
                "",
                "",
                deviceType,
            )
            return True

    else:
        printReport(conn, host, f"Test 17 successfull for {host} - n/a", "", "", deviceType)
        return True


def printReport(conn, host, outcome, errors, commands, deviceType):
    # Function that writes the errors found in the different tests to a file and stores it in the folder REPORTS. It doesn't return anything.
    if errors and commands:
        if os.path.exists("REPORTS") == False:
            os.mkdir("REPORTS")

        with open("REPORTS/" + host + "-Report.txt", "a") as report:

            report.write("ISSUE FOUND: " + str(outcome) + "\n")

            if deviceType == "arista_eos":  # Arista

                for n in errors.keys():
                    report.write("\nISSUE DETAIL: " + str(errors[n]) + "\n")
                    report.write("\n# Detailed information #\n")
                    for i in commands[n].keys():

                        try:

                            output = conn.enable(commands[n][i], encoding="text")

                        except:
                            report.write(
                                "\nWARNING: The command '"
                                + commands[n][i]
                                + "' failed for the device "
                                + host
                                + ". Check the command syntax or run it manually on the device\n"
                            )
                            continue

                        report.write(host + "#" + str(output) + "\n")
                        report.write(str(output[0]["result"]["output"]) + "\n")
                        updateMSExcelReport(
                            host,
                            outcome,
                            errors[n],
                            host + "#" + str(output[0]["command"]) + "\n" + str(output[0]["result"]["output"]),
                        )

            if deviceType == "juniper_junos":  # Juniper
                for n in errors.keys():
                    report.write("\nISSUE DETAIL: " + str(errors[n]) + "\n")
                    report.write("\n# Detailed information #\n")
                    for i in commands[n].keys():

                        try:
                            output = getDeviceOutput(host, commands[n][i])
                        except:
                            report.write(
                                "\nWARNING: The command '"
                                + commands[n][i]
                                + "' failed for the device "
                                + host
                                + ". Check the command syntax or run it manually on the device\n"
                            )
                            continue

                        report.write(host + ">" + str(commands[n][i]) + "\n")
                        report.write(str(output) + "\n")
                        updateMSExcelReport(
                            host, outcome, errors[n], host + "#" + str(commands[n][i]) + "\n" + str(output)
                        )

            if deviceType == "sr_linux":  # Nokia

                for n in errors.keys():
                    report.write("\nISSUE DETAIL: " + str(errors[n]) + "\n")
                    report.write("\n# Detailed information #\n")
                    for i in commands[n].keys():
                        textOutput = False
                        try:
                            if commands[n][i].startswith("show"):
                                output = getDeviceOutput(host, commands[n][i])
                                textOutput = True
                            else:
                                output = getNokiaInformation(host, commands[n][i])

                        except:
                            report.write(
                                "\nWARNING: The command '"
                                + commands[n][i]
                                + "' failed for the device "
                                + host
                                + ". Check the command syntax or run it manually on the device\n"
                            )
                            continue

                        report.write(host + "#" + str(commands[n][i]) + "\n")
                        if not textOutput:
                            pprint.pprint(output, stream=report)
                            report.write(str(pprint.pformat(output)) + "\n")
                            updateMSExcelReport(
                                host,
                                outcome,
                                errors[n],
                                host + "#" + str(commands[n][i]) + "\n" + str(pprint.pformat(output)),
                            )
                        else:
                            updateMSExcelReport(
                                host, outcome, errors[n], host + "#" + str(commands[n][i]) + "\n" + str(output)
                            )
                            report.write(str(output) + "\n")

            if deviceType == "other":
                for n in errors.keys():
                    report.write("\nISSUE DETAIL: " + str(errors[n]) + "\n")
                    updateMSExcelReport(host, outcome, errors[n], "")

    else:
        updateMSExcelReport(host, outcome, "", "")

    return


def createMSExcelReport(hosts, deviceTypes):
    if os.path.exists("EXCELREPORTS") == False:
        os.mkdir("EXCELREPORTS")
    for i in range(len(hosts)):

        # for host, deviceType in hosts, deviceTypes:
        if os.path.exists("EXCELREPORTS/" + str(hosts[i]) + ".xlsx") == False:

            wb = Workbook()

            wb.remove(wb["Sheet"])
            wb.create_sheet(index=1, title=hosts[i])
            ws1 = wb[hosts[i]]
            if str(deviceTypes[i]) == "s_linux":
                thin = Side(border_style="thin", color="000000")
                ws1["B2"] = "Checklist"
                ws1["B2"].font = Font(name="Arial", size=11, bold=True)
                ws1["B2"].fill = PatternFill("solid", fgColor="7E92E5")
                ws1["B2"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["B2"].alignment = Alignment(horizontal="center", vertical="center")
                ws1["C2"] = "Status"
                ws1["C2"].font = Font(name="Arial", size=11, bold=True)
                ws1["C2"].fill = PatternFill("solid", fgColor="7E92E5")
                ws1["C2"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["C2"].alignment = Alignment(horizontal="center", vertical="center")
                ws1["D2"] = "Remark"
                ws1["D2"].font = Font(name="Arial", size=11, bold=True)
                ws1["D2"].fill = PatternFill("solid", fgColor="7E92E5")
                ws1["D2"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["D2"].alignment = Alignment(horizontal="center", vertical="center")
                ws1["E2"] = "Checked by"
                ws1["E2"].font = Font(name="Arial", size=11, bold=True)
                ws1["E2"].fill = PatternFill("solid", fgColor="7E92E5")
                ws1["E2"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["E2"].alignment = Alignment(horizontal="center", vertical="center")
                ws1["B3"] = "Test 1 RANCID tag in Netbox:"
                ws1["B3"].font = Font(name="Arial", size=11, bold=True)
                ws1["B3"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["B3"].alignment = Alignment(horizontal="justify", vertical="center")
                ws1["B4"] = "Test 2 Check the Firmware against Netbox: "
                ws1["B4"].font = Font(name="Arial", size=11, bold=True)
                ws1["B4"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["B4"].alignment = Alignment(horizontal="justify", vertical="center")
                ws1["B5"] = "Test 3 Check device Serial against Netbox: "
                ws1["B5"].font = Font(name="Arial", size=11, bold=True)
                ws1["B5"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["B5"].alignment = Alignment(horizontal="justify", vertical="center")
                ws1["B6"] = "Test 4 Check device Model against Netbox: "
                ws1["B6"].font = Font(name="Arial", size=11, bold=True)
                ws1["B6"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["B6"].alignment = Alignment(horizontal="justify", vertical="center")
                ws1["B7"] = "Test 5 Check MAC against Netbox: "
                ws1["B7"].font = Font(name="Arial", size=11, bold=True)
                ws1["B7"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["B7"].alignment = Alignment(horizontal="justify", vertical="center")
                ws1["B8"] = "Test 6 Check primary IPv4 against Netbox: "
                ws1["B8"].font = Font(name="Arial", size=11, bold=True)
                ws1["B8"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["B8"].alignment = Alignment(horizontal="justify", vertical="center")
                ws1["B9"] = "Test 7 Check Status of Power Outlets:"
                ws1["B9"].font = Font(name="Arial", size=11, bold=True)
                ws1["B9"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["B9"].alignment = Alignment(horizontal="justify", vertical="center")
                wb.save("EXCELREPORTS/" + str(hosts[i]) + ".xlsx")

            elif str(deviceTypes[i]) == "o_linux":
                thin = Side(border_style="thin", color="000000")
                ws1["B2"] = "Checklist"
                ws1["B2"].font = Font(name="Arial", size=11, bold=True)
                ws1["B2"].fill = PatternFill("solid", fgColor="7E92E5")
                ws1["B2"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["B2"].alignment = Alignment(horizontal="center", vertical="center")
                ws1["C2"] = "Status"
                ws1["C2"].font = Font(name="Arial", size=11, bold=True)
                ws1["C2"].fill = PatternFill("solid", fgColor="7E92E5")
                ws1["C2"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["C2"].alignment = Alignment(horizontal="center", vertical="center")
                ws1["D2"] = "Remark"
                ws1["D2"].font = Font(name="Arial", size=11, bold=True)
                ws1["D2"].fill = PatternFill("solid", fgColor="7E92E5")
                ws1["D2"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["D2"].alignment = Alignment(horizontal="center", vertical="center")
                ws1["E2"] = "Checked by"
                ws1["E2"].font = Font(name="Arial", size=11, bold=True)
                ws1["E2"].fill = PatternFill("solid", fgColor="7E92E5")
                ws1["E2"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["E2"].alignment = Alignment(horizontal="center", vertical="center")
                ws1["B3"] = "Test 1 RANCID tag in Netbox:"
                ws1["B3"].font = Font(name="Arial", size=11, bold=True)
                ws1["B3"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["B3"].alignment = Alignment(horizontal="justify", vertical="center")
                ws1["B4"] = "Test 2 Check Device Model against Netbox:"
                ws1["B4"].font = Font(name="Arial", size=11, bold=True)
                ws1["B4"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["B4"].alignment = Alignment(horizontal="justify", vertical="center")
                ws1["B5"] = "Test 3 Check Primary IPv4 against Netbox:"
                ws1["B5"].font = Font(name="Arial", size=11, bold=True)
                ws1["B5"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["B5"].alignment = Alignment(horizontal="justify", vertical="center")
                ws1["B6"] = "Test 4 Check device Serial against Netbox:"
                ws1["B6"].font = Font(name="Arial", size=11, bold=True)
                ws1["B6"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["B6"].alignment = Alignment(horizontal="justify", vertical="center")
                ws1["B7"] = "Test 5 Check Host Name against Netbox:"
                ws1["B7"].font = Font(name="Arial", size=11, bold=True)
                ws1["B7"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["B7"].alignment = Alignment(horizontal="justify", vertical="center")
                ws1["B8"] = "Test 6 Check device Firmware against Netbox:"
                ws1["B8"].font = Font(name="Arial", size=11, bold=True)
                ws1["B8"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["B8"].alignment = Alignment(horizontal="justify", vertical="center")
                ws1["B9"] = "Test 7 Check Console port labels against Netbox:"
                ws1["B9"].font = Font(name="Arial", size=11, bold=True)
                ws1["B9"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["B9"].alignment = Alignment(horizontal="justify", vertical="center")
                ws1["B10"] = "Test 8 Check devices connected to Console ports against Console Labels:"
                ws1["B10"].font = Font(name="Arial", size=11, bold=True)
                ws1["B10"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["B10"].alignment = Alignment(horizontal="justify", vertical="center")
                wb.save("EXCELREPORTS/" + str(hosts[i]) + ".xlsx")

            else:
                # Applying styles to the template
                thin = Side(border_style="thin", color="000000")
                ws1["B2"] = "Checklist"
                ws1["B2"].font = Font(name="Arial", size=11, bold=True)
                ws1["B2"].fill = PatternFill("solid", fgColor="7E92E5")
                ws1["B2"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["B2"].alignment = Alignment(horizontal="center", vertical="center")
                ws1["B3"] = "Test 1: Check software versions"
                ws1["B3"].font = Font(name="Arial", size=11, bold=True)
                ws1["B3"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["B3"].alignment = Alignment(horizontal="justify", vertical="center")
                ws1["B4"] = "Test 2: Check physical serial numbers againts netbox"
                ws1["B4"].font = Font(name="Arial", size=11, bold=True)
                ws1["B4"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["B4"].alignment = Alignment(horizontal="justify", vertical="center")
                ws1["B5"] = "Test 3: Check physical connections and neighbors againts Netbox"
                ws1["B5"].font = Font(name="Arial", size=11, bold=True)
                ws1["B5"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["B5"].alignment = Alignment(horizontal="justify", vertical="center")
                ws1["B6"] = "Test 4: Check physical wiring againts Netbox and interface status UP"
                ws1["B6"].font = Font(name="Arial", size=11, bold=True)
                ws1["B6"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["B6"].alignment = Alignment(horizontal="justify", vertical="center")
                ws1["B7"] = "Test 5: Check link light levels"
                ws1["B7"].font = Font(name="Arial", size=11, bold=True)
                ws1["B7"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["B7"].alignment = Alignment(horizontal="justify", vertical="center")
                ws1["B8"] = "Test 6: Check logs for bouncing Interfaces"
                ws1["B8"].font = Font(name="Arial", size=11, bold=True)
                ws1["B8"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["B8"].alignment = Alignment(horizontal="justify", vertical="center")
                ws1["B9"] = "Test 7: Check environment (PSUs, temperature sensors...)"
                ws1["B9"].font = Font(name="Arial", size=11, bold=True)
                ws1["B9"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["B9"].alignment = Alignment(horizontal="justify", vertical="center")
                ws1["B10"] = "Test 8: Confirm BGP v4/v6 and BGP EVPN neighbors UP"
                ws1["B10"].font = Font(name="Arial", size=11, bold=True)
                ws1["B10"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["B10"].alignment = Alignment(horizontal="justify", vertical="center")
                ws1["B11"] = "Test 9: CRC Ping flood check for link IP and interface errors"
                ws1["B11"].font = Font(name="Arial", size=11, bold=True)
                ws1["B11"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["B11"].alignment = Alignment(horizontal="justify", vertical="center")
                ws1["B12"] = "Test 10: Check Arista devices have aliases configured"
                ws1["B12"].font = Font(name="Arial", size=11, bold=True)
                ws1["B12"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["B12"].alignment = Alignment(horizontal="justify", vertical="center")
                ws1["B13"] = "Test 11: Check lo2 is configured and matches the Netbox information"
                ws1["B13"].font = Font(name="Arial", size=11, bold=True)
                ws1["B13"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["B13"].alignment = Alignment(horizontal="justify", vertical="center")
                ws1["B14"] = "Test 12: Checks sflow on transit/pni/ix conections and jumbo MTU values"
                ws1["B14"].font = Font(name="Arial", size=11, bold=True)
                ws1["B14"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["B14"].alignment = Alignment(horizontal="justify", vertical="center")
                ws1["B15"] = "Test 13: Check NH for the default route on vrf PACKET INTERNAL in control racks"
                ws1["B15"].font = Font(name="Arial", size=11, bold=True)
                ws1["B15"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["B15"].alignment = Alignment(horizontal="justify", vertical="center")
                ws1["B16"] = "Test 14: RANCID tag in Netbox"
                ws1["B16"].font = Font(name="Arial", size=11, bold=True)
                ws1["B16"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["B16"].alignment = Alignment(horizontal="justify", vertical="center")
                ws1["B17"] = "Test 15: Packetbot User configured on Nokia/Arista devices"
                ws1["B17"].font = Font(name="Arial", size=11, bold=True)
                ws1["B17"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["B17"].alignment = Alignment(horizontal="justify", vertical="center")
                ws1["B18"] = "Test 16: Device hostname in Net-Utils /etc/hosts"
                ws1["B18"].font = Font(name="Arial", size=11, bold=True)
                ws1["B18"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["B18"].alignment = Alignment(horizontal="justify", vertical="center")
                ws1["B19"] = "Test 17: Check DHCP Forwarding is correct in MSW"
                ws1["B19"].font = Font(name="Arial", size=11, bold=True)
                ws1["B19"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["B19"].alignment = Alignment(horizontal="justify", vertical="center")
                ws1["C2"] = "Status"
                ws1["C2"].font = Font(name="Arial", size=11, bold=True)
                ws1["C2"].fill = PatternFill("solid", fgColor="7E92E5")
                ws1["C2"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["C2"].alignment = Alignment(horizontal="center", vertical="center")
                ws1["D2"] = "Remark"
                ws1["D2"].font = Font(name="Arial", size=11, bold=True)
                ws1["D2"].fill = PatternFill("solid", fgColor="7E92E5")
                ws1["D2"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["D2"].alignment = Alignment(horizontal="center", vertical="center")
                ws1["E2"] = "Checked by"
                ws1["E2"].font = Font(name="Arial", size=11, bold=True)
                ws1["E2"].fill = PatternFill("solid", fgColor="7E92E5")
                ws1["E2"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws1["E2"].alignment = Alignment(horizontal="center", vertical="center")
                wb.save("EXCELREPORTS/" + str(hosts[i]) + ".xlsx")

    return


def updateMSExcelReport(host, outcome, error, commandOutput):
    # Function that works along with printReport() function and creates the a report in Microsoft Excel format that can be imported easily into Google sheets.

    wb = openpyxl.load_workbook("EXCELREPORTS/" + str(host) + ".xlsx")
    thin = Side(border_style="thin", color="000000")
    ws = wb[host]
    for testNumbr in range(1, 20):
        if str("TEST") + str(testNumbr) == str(outcome.split()[0].upper()) + str(outcome.split()[1].upper()):
            if not error:
                ws["C" + str(testNumbr + 2)].font = Font(name="Arial", size=11, color="00339966")
                ws["C" + str(testNumbr + 2)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws["C" + str(testNumbr + 2)].alignment = Alignment(horizontal="center", vertical="center")
                if "n/a" in outcome:
                    ws["C" + str(testNumbr + 2)] = "N/A"
                else:
                    ws["C" + str(testNumbr + 2)] = "PASSED"
                ws["D" + str(testNumbr + 2)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws["D" + str(testNumbr + 2)].alignment = Alignment(horizontal="justify", vertical="center")
                ws["E" + str(testNumbr + 2)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws["E" + str(testNumbr + 2)].alignment = Alignment(horizontal="justify", vertical="center")
                ws["E" + str(testNumbr + 2)] = "Checked by post-checks script"
            else:
                ws["C" + str(testNumbr + 2)].font = Font(name="Arial", size=11, color="00FF0000")
                ws["C" + str(testNumbr + 2)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws["C" + str(testNumbr + 2)].alignment = Alignment(horizontal="center", vertical="center")
                ws["C" + str(testNumbr + 2)] = "FAILED"
                ws["D" + str(testNumbr + 2)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws["D" + str(testNumbr + 2)].alignment = Alignment(horizontal="justify", vertical="center")
                old_cell_value = ws["D" + str(testNumbr + 2)].value if ws["D" + str(testNumbr + 2) + "\n"].value else ""
                ws["D" + str(testNumbr + 2)].value = str(old_cell_value) + str(error) + "\n" + str(commandOutput)
                ws["E" + str(testNumbr + 2)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws["E" + str(testNumbr + 2)].alignment = Alignment(horizontal="justify", vertical="center")
                ws["E" + str(testNumbr + 2)] = "Checked by post-checks script"

    wb.save("EXCELREPORTS/" + str(host) + ".xlsx")
    return


def checkReachability(hostname, deviceType):
    # Function that checks if the devices are reachable from the server where the script is running. If device is not reachable, the script won't try to establish a connection with the device API.
    # Returns True if the devices re reachable and False if not.
    if deviceType == "s_linux" or deviceType == "o_linux":
        deviceType = "other"
    if sys.platform[:3] == "win":
        response = os.system("ping -n 1 " + hostname + " > NUL")
    else:
        response = os.system("ping -c 1 " + hostname + " >/dev/null")

    if response == 0:
        return True
    else:
        print("\nWARNING: The device " + hostname + " is unreachable. No checks will be made.")
        errors = {"1": f"WARNING: The device {hostname} is unreachable. No checks will be made."}
        printReport("", hostname, "Test 1 failed for " + hostname, errors, f"ping {hostname}", deviceType)
        return False


def checkAristaAliases(conn, host, deviceType):
    # Function that checks that all the alias a configured on the Arista devices. If all the alias are configured it returns True and if not,
    # it returns False and prints the missing alias in the report
    commandsReport = {}
    currentAliases = ["senz", "srnz", "rebuild-copp", "conint", "spd", "sqnz", "shmc", "snz"]
    mswCurrentAliases = ["senz", "srnz", "rebuild-copp", "conint", "spd", "sqnz", "shmc", "snz", "rebuild-ipmi-acl"]
    errors = {}
    i = 1
    if deviceType == "arista_eos":
        command = "show aliases"
        try:
            output = conn.enable(command)

        except:
            print(
                "\nWARNING: The command "
                + command
                + " failed for the device "
                + host
                + ". Check the command syntax or run it manually on the device. Skipping Test 2 ...\n"
            )
            return

        for alias in output[0]["result"]["aliases"]:

            if (
                alias not in currentAliases and host.find("msw") == -1
            ):  # WMSWs are skipped because they have one more alias than the rest pof the Arista devices
                errors[i] = "The device " + host + " doesn't have the following alias: " + alias
                commandsReport[i] = {1: command}
                i += 1

            if host.find("msw") != -1 and alias not in mswCurrentAliases:  # Check for the MSWs
                errors[i] = "The device " + host + " doesn't have the following alias: " + alias
                commandsReport[i] = {1: command}
                i += 1

        if errors:
            printReport(
                conn,
                host,
                "Test 10 failed for " + host + " - The device has one or more missing aliases",
                errors,
                commandsReport,
                deviceType,
            )
            return False

        else:

            # print("Test 10 successfull for "+host+" - The device has all the required aliases")
            printReport(
                conn,
                host,
                "Test 10 successfull for " + host + " - The device has all the required aliases",
                "",
                "",
                deviceType,
            )
            return True

    else:
        printReport(conn, host, "Test 10 successfull for " + host + " - n/a", "", "", deviceType)
        return True


def createNewWorkbook(tmpWb, outputWB):
    # Function used by the mergeMSExcelReports function to create the Post-checks excel file
    for wb in tmpWb:
        for sheetName in wb.sheetnames:
            o = outputWB.create_sheet(sheetName)
            safeTitle = o.title
            copySheet(wb[sheetName], outputWB[safeTitle])


def copySheet(sourceSheet, newSheet):
    # Function used by the mergeMSExcelReports function to copy the different sheets to the Post-checks excel file
    for row in sourceSheet.rows:
        for cell in row:
            newCell = newSheet.cell(row=cell.row, column=cell.col_idx, value=cell.value)
            if cell.has_style:
                newCell.font = copy(cell.font)
                newCell.border = copy(cell.border)
                newCell.fill = copy(cell.fill)
                newCell.number_format = copy(cell.number_format)
                newCell.protection = copy(cell.protection)
                newCell.alignment = copy(cell.alignment)


def mergeMSExcelReports():
    # Function that reads all the MS Excel files contianed in the folder EXCELREPORTS and consolidates all info under one file called Post-checks.xlsx created in the post-checks folder.
    # It also cleans the TMP files of the EXCELREPORTS
    inputFiles = os.listdir("EXCELREPORTS/")
    outputFile = "Post-checks.xlsx"
    tmpWBFiles = [load_workbook("EXCELREPORTS/" + str(f)) for f in inputFiles]

    outputWB = Workbook()
    del outputWB["Sheet"]  # We want our new book to be empty
    createNewWorkbook(tmpWBFiles, outputWB)
    outputWB.save(outputFile)
    return True


def checkLo2IPAddresses(conn, host, deviceType):
    commandsReport = {}
    errors = {}
    peerLo2IPAddress = ""
    hostLo2IPAddress = ""
    skipCheck = False
    peerHost = ""
    i = 1
    if deviceType == "arista_eos" and (
        host.find("esr") != -1 or host.find("csr") != -1
    ):  # Check if device trype is Arista and if it is an esr or csrs

        # Get the peer switch hostname and its lo2 ip address
        if host.find("esr") != -1:
            if host[4] == "a":

                peerHost = host.split(".")[0].replace("a", "b")
            else:

                peerHost = host.split(".")[0].replace("b", "a")

            for numChar in range(1, len(host.split("."))):
                peerHost = str(peerHost) + "." + str(host.split(".")[numChar])

                # Check Looback2 ip addresses configured on Netbox #
        hostLoppbacksNB = getDevLoopbacksNB(host)
        if host.find("esr") != -1:
            peerLoopbacksNB = getDevLoopbacksNB(peerHost)
            for n in hostLoppbacksNB.keys():
                if hostLoppbacksNB[n]["interface"] == "loopback2" and hostLoppbacksNB[n]["addr"]:  # Only Lo2s
                    hostLo2IPAddress = hostLoppbacksNB[n]["addr"]
            for m in peerLoopbacksNB.keys():
                if peerLoopbacksNB[m]["interface"] == "loopback2" and peerLoopbacksNB[m]["addr"]:  # Only Lo2s
                    peerLo2IPAddress = peerLoopbacksNB[m]["addr"]

            if peerLo2IPAddress and hostLo2IPAddress:
                if peerLo2IPAddress != hostLo2IPAddress:
                    errors[i] = (
                        "The peers devices "
                        + host
                        + " and "
                        + peerHost
                        + " have different ip addresses for the loopback2 in Netbox"
                    )
                    commandsReport[i] = {1: "show interfaces loopback2"}
                    i = i + 1
            else:
                errors[i] = (
                    " The host "
                    + host
                    + " or "
                    + peerHost
                    + " doesn't have IP addresses assigned for the loopback2 in Netbox"
                )
                commandsReport[i] = {1: "show interfaces loopback2"}
                i = i + 1
                skipCheck = True
        elif host.find("csr") != -1:
            for n in hostLoppbacksNB.keys():
                if hostLoppbacksNB[n]["interface"] == "loopback2" and hostLoppbacksNB[n]["addr"]:  # Only Lo2s
                    hostLo2IPAddress = hostLoppbacksNB[n]["addr"]

        # Getting the lo2 ip address from the switch
        command = "show interfaces loopback2"
        try:
            output = conn.enable(command)

        except:
            print(
                "\nWARNING: The command "
                + command
                + " failed for the device "
                + host
                + ". Check the command syntax or run it manually on the device. Skipping Test 11 ...\n"
            )
            return False

        # Compare lo2 ip address of peer host matches the one that is configured on the device
        if output[0]["result"]["interfaces"]["Loopback2"]["interfaceAddress"]:
            if host.find("esr") != -1:  # ESRs
                if (
                    peerLo2IPAddress.split("/")[0]
                    != output[0]["result"]["interfaces"]["Loopback2"]["interfaceAddress"][0]["primaryIp"]["address"]
                ):
                    errors[i] = (
                        "The peers devices "
                        + host
                        + " and "
                        + peerHost
                        + " have different IP addresses configured for the loopback2. The Peer Lo2 IP address is "
                        + str(peerLo2IPAddress.split("/")[0])
                        + " and the device Lo2 IP address is "
                        + str(
                            output[0]["result"]["interfaces"]["Loopback2"]["interfaceAddress"][0]["primaryIp"][
                                "address"
                            ]
                        )
                    )
                    commandsReport[i] = {1: command}
                    i = i + 1
            elif host.find("csr") != -1:  # CSRs
                if (
                    hostLo2IPAddress.split("/")[0]
                    != output[0]["result"]["interfaces"]["Loopback2"]["interfaceAddress"][0]["primaryIp"]["address"]
                ):
                    errors[i] = (
                        "The "
                        + host
                        + " configured Lo2 IP address and the Netbox Lo2 IP address are different. Netbox Lo2 IP address is "
                        + str(hostLo2IPAddress.split("/")[0])
                        + "and the cofnigured Lo2 IP address is "
                        + str(
                            output[0]["result"]["interfaces"]["Loopback2"]["interfaceAddress"][0]["primaryIp"][
                                "address"
                            ]
                        )
                    )
                    commandsReport[i] = {1: command}
                    i = i + 1
                    skipCheck = True

        else:  # No interface ip address configured
            errors[i] = "The device " + host + " doesn't have the Loopback 2 IP address configured"
            commandsReport[i] = {1: command}
            i = i + 1
            skipCheck = True

        # Check to confirm that loopback ip addresses are included in the prefix-list (only for csrs)
        if host.find("csr") != -1 and not skipCheck:  # ONLY FOR CSRs
            command = "show ip prefix-list LOOPBACKS"
            try:
                outputPefixList = conn.enable(command)

            except:
                print(
                    "\nWARNING: The command "
                    + command
                    + " failed for the device "
                    + host
                    + ". Check the command syntax or run it manually on the device. Skipping Test 11 ...\n"
                )
                return False

            match = False
            for entry in outputPefixList[0]["result"]["ipPrefixLists"]["LOOPBACKS"]["ipPrefixEntries"]:
                if entry["prefix"] == str(
                    output[0]["result"]["interfaces"]["Loopback2"]["interfaceAddress"][0]["primaryIp"]["address"]
                ) + "/" + str(
                    output[0]["result"]["interfaces"]["Loopback2"]["interfaceAddress"][0]["primaryIp"]["maskLen"]
                ):
                    match = True

            if not match:
                errors[i] = (
                    "The device "
                    + host
                    + " loopback2 IP address is not configured on the prefix-list LOOPBACKS. The Lo2 ip address is "
                    + str(output[0]["result"]["interfaces"]["Loopback2"]["interfaceAddress"][0]["primaryIp"]["address"])
                    + "/"
                    + str(output[0]["result"]["interfaces"]["Loopback2"]["interfaceAddress"][0]["primaryIp"]["maskLen"])
                )
                commandsReport[i] = {1: command}
                i = i + 1

        if errors:
            if host.find("esr") != -1:
                printReport(
                    conn,
                    host,
                    "Test 11 failed for "
                    + host
                    + " - This device and its peer switch "
                    + peerHost
                    + " have different ip addresses for the Loopback2\n",
                    errors,
                    commandsReport,
                    deviceType,
                )

            elif host.find("csr") != -1:
                printReport(
                    conn,
                    host,
                    "Test 11 failed for "
                    + host
                    + " - Either the Loopback 2 IP address is not configured preoperly or the Loopback 2 IP address is not configured in the prefix-list LOOPBACKS.\n",
                    errors,
                    commandsReport,
                    deviceType,
                )

            return False

        else:

            # print("Test 11 successfull both "+host+" and "+ peerHost+" have the same IP address for the Loopback2 (in Netbox and configured)"
            printReport(
                conn,
                host,
                "Test 11 successfull both "
                + host
                + " and "
                + peerHost
                + " have the same IP address for the Loopback2 (in Netbox and configured)",
                "",
                "",
                deviceType,
            )
            return True

    else:
        printReport(conn, host, "Test 11 successfull - n/a)", "", "", deviceType)
        return True


def checkSflowAndMTU(conn, host, deviceType):
    # Function that checks if TRANSIT, PNI and PEERING connection has SFLOW enabled and MTU value of 1514
    if deviceType == "juniper_junos":  # Juniper
        if "fw" not in host:
            commandsReport = {}
            errors = {}
            i = 1
            mtu_val = "1514"
            target_ifaces = {}
            conf_ifaces = conn.rpc.get_interface_information(extensive=True)  # get all intefaces from the device
            for iface in conf_ifaces.findall(
                ".//description"
            ):  # filter interfaces based on the description, get a list of 'interesting' interfaces
                name = iface.find("..//name").text
                try:
                    descr = iface.find("../description").text.split("{")[0].split(": ")[1]
                except (IndexError, ValueError):
                    descr = iface.find("../description").text
                if "TRANSIT" in iface.text or "PNI" in iface.text or "PEERING" in iface.text:
                    if not "OOB" in iface.text:
                        if "ae" not in name:
                            target_ifaces.update(
                                {name.strip("\n"): descr.strip("\n")}
                            )  # get the list of interfaces to check
            for key, value in target_ifaces.items():
                command = "show sflow interface interface-name " + str(key)
                try:
                    sflowconf = conn.rpc.get_sflow_interface(interface_name=key).find(
                        ".//interface-status-egress"
                    )  # if the interface doesn't have sflow this value will be None
                except:
                    print(
                        "\nWARNING: The command '"
                        + command
                        + "' failed for the device "
                        + host
                        + ". Check the command syntax or run it manually on the device. Skipping the test of the following interface: "
                        + str(key)
                    )
                if sflowconf == None:
                    errors[i] = (
                        "The device "
                        + host
                        + " doesn't have SFLOW enabled for the interface "
                        + str(key)
                        + " ("
                        + str(value)
                        + ")"
                    )
                    commandsReport[i] = {1: command}
                    i = i + 1
                mtu = (
                    conn.rpc.get_interface_information(extensive=True, interface_name=str(key))
                    .find(".//mtu")
                    .text.strip("\n")
                )
                command = "show interface " + str(key) + " extensive"
                if mtu != mtu_val:
                    errors[i] = (
                        "The device " + host + " interface " + str(key) + " (" + str(value) + ") has Jumbo MTU enabled"
                    )
                    commandsReport[i] = {1: command}
                    i = i + 1
            if errors:
                printReport(
                    conn,
                    host,
                    "Test 12 failed for "
                    + host
                    + " - There is at least one Transit/IX/PNI interface without SFLOW enabled or with wrong MTU value",
                    errors,
                    commandsReport,
                    deviceType,
                )
                return False
            else:
                printReport(
                    conn,
                    host,
                    "Test 12 successfull for "
                    + host
                    + " - There aren't Transit/IX/PNI interfaces without sflow and wrong MTU values set",
                    "",
                    "",
                    deviceType,
                )
                return True
    else:
        printReport(conn, host, "Test 12 successfull for " + host + " - n/a", "", "", deviceType)
        return True


def checkNHdefault(conn, host, deviceType):
    # This function get the NH address for the default route on CR TOR and check them to be towards BBR
    commandsReport = {}
    errors = {}
    i = 1
    devtags = NBgetTags(host)
    if deviceType == "arista_eos" and host.find("esr") != -1 and "packet-internal" in devtags:
        commands = ["show ip route vrf PACKET-INTERNAL 0/0", "show ipv6 route vrf PACKET-INTERNAL ::/0"]
        NextHops = []
        for command in commands:
            try:
                output = conn.enable(command)

            except:
                print(
                    "\nWARNING: The command '"
                    + command
                    + "' failed for the device "
                    + host
                    + ". Check the command syntax or run it manually on the device. Skipping this command ...\n"
                )
                continue
            nh = 0
            if command.find("ipv6") == -1:
                while nh < len(output[0]["result"]["vrfs"]["PACKET-INTERNAL"]["routes"]["0.0.0.0/0"]["vias"]):
                    if (
                        output[0]["result"]["vrfs"]["PACKET-INTERNAL"]["routes"]["0.0.0.0/0"]["vias"][nh]["vtepAddr"]
                        not in NextHops
                    ):
                        NextHops.append(
                            output[0]["result"]["vrfs"]["PACKET-INTERNAL"]["routes"]["0.0.0.0/0"]["vias"][nh][
                                "vtepAddr"
                            ]
                        )
                    nh += 1
            else:
                while nh < len(output[0]["result"]["routes"]["::/0"]["vias"]):
                    if output[0]["result"]["routes"]["::/0"]["vias"][nh]["vtepAddr"] not in NextHops:
                        NextHops.append(output[0]["result"]["routes"]["::/0"]["vias"][nh]["vtepAddr"])
                    nh += 1
        nh = 0
        while nh < len(NextHops):
            devName = NBgetNameFromIP(NextHops[nh])
            nh += 1
            if devName.find("bbr") == -1:
                errors[i] = (
                    "The device "
                    + host
                    + " doesn't have BBR's Lo0 as next hop for the default route on the PACKET-INTERNAL vrf"
                )
                commandsReport[i] = {
                    1: "show ip route vrf PACKET-INTERNAL 0/0",
                    2: "show ipv6 route vrf PACKET-INTERNAL ::/0",
                }
                i += 1
        if errors:
            printReport(
                conn,
                host,
                "Test 13 failed for "
                + host
                + " - There is at least one default route on the PACKET-INTERNAL vrf whose next hop is not the BBR",
                errors,
                commandsReport,
                deviceType,
            )
            return False
        else:
            printReport(
                conn,
                host,
                "Test 13 successfull for " + host + " - The control rack default route NH are correct",
                "",
                "",
                deviceType,
            )
            return True
    else:
        printReport(conn, host, "Test 13 successfull for " + host + " - n/a", "", "", deviceType)
        return True


def checkRancidTag(conn, host, deviceType):
    # This function checks if the right device have the rancid tag
    commandsReport = {}
    errors = {}
    devtags = NBgetTags(host)
    i = 1
    dev_list = {
        "bbr",
        "bsr",
        "csr",
        "dsr",
        "esr",
        "fw",
        "mdr",
        "mrr",
        "msr",
        "ssp",
        "msw",
        "cs0",
        "cs1",
        "cs2",
        "cs3",
        "cs4",
    }
    DeviceName = host[:3]
    if DeviceName.startswith(("cs0", "cs1", "cs2", "cs3", "cs4")) or DeviceName.startswith("pdu"):
        testnum = "Test 1"
    else:
        testnum = "Test 14"

    if DeviceName.lower() in dev_list and "rancid" not in devtags:
        errors[i] = "The device " + host + " doesn't have rancid tag"
        commandsReport[i] = {1: "Rancid tag is missing"}
        i += 1
        updateRancid(host, "Add", testnum, conn)
    elif DeviceName.lower() not in dev_list and "rancid" in devtags:
        errors[i] = "Please remove the rancid tag for " + host
        commandsReport[i] = {1: "Need to remove rancid tag"}
        i += 1
        updateRancid(host, "Remove", testnum, conn)
    else:
        printReport(
            conn, host, f"{testnum} successfull for {host} - The rancid tag is present or not required", "", "", "other"
        )
        return True

    if errors:
        printReport(
            conn,
            host,
            f"{testnum} failed for {host} - there are issues with the rancid tag. ",
            errors,
            commandsReport,
            "other",
        )
        return False
    else:
        printReport(conn, host, f"{testnum} successfull for {host} - The rancid tag is present", "", "", "other")
        return True


def updateRancid(host, action, testnum, conn):
    nb = pynetbox.api(url=os.environ["NB_URL"], token=os.environ["NB_API_KEY"])
    device = nb.dcim.devices.filter(name=host)

    # Adding a tag to the device by tag ID
    tag_to_add = nb.extras.tags.get(id=18)

    for tags in device:
        if action == "Add":
            tags.tags.append(tag_to_add)
            errors = {"Tag added": f"Rancid tag is added to the device {host}"}
            printReport(conn, host, f"{testnum} failed for {host}", errors, " ", "other")
        else:
            tags.tags.remove(tag_to_add)
            errors = {"Tag removed": f"Rancid tag is removed from the device {host}"}
            printReport(conn, host, f"{testnum} failed for {host}", errors, " ", "other")
        tags.save()


def checkPacketbotUser(conn, host, deviceType):
    # Function that checks if the packetbot user is configured on Nokia devices. It returns True if the user is configured and False if it doesn't
    commandsReport = {}
    errors = {}
    i = 1
    if re.findall(r"csr|esr|msw", host, re.I):
        if deviceType == "sr_linux" or deviceType == "arista_eos":  # Nokia/Arista

            command = "/system/aaa/authentication/user" if deviceType == "sr_linux" else "/system/aaa"
            try:
                userInfo = getNokiaInformation(host, command)

            except:
                print(
                    "\nWARNING: The command "
                    + command
                    + " failed for the device "
                    + host
                    + ". Check the command syntax or run it manually on the device. Skipping Test 6 ...\n"
                )
                return

            userFound = False

            if deviceType == "arista_eos":
                users = userInfo["notification"][0]["update"][0]["val"]["openconfig-system:authentication"]["users"][
                    "user"
                ]
            else:
                users = userInfo["notification"][0]["update"][0]["val"]["user"]

            for user in users:
                if user["username"].lower() == "packetbot":
                    userFound = True
                    break

            if not userFound:
                errors[i] = "The device " + host + " doesn't have the packetbot user configured"
                commandsReport[i] = {1: command}
                i += 1

        if errors:
            printReport(
                conn,
                host,
                "Test 15 failed for " + host + " - The packetbot user is not configured on this Nokia/Arista device. ",
                errors,
                commandsReport,
                deviceType,
            )
            return False
        else:
            printReport(
                conn,
                host,
                "Test 15 successfull for " + host + " - The packetbot user is configured on the device.",
                "",
                "",
                deviceType,
            )
            return True

    else:
        printReport(conn, host, "Test 15 successfull for " + host + " - n/a", "", "", deviceType)
        return True


def checkHostsFiles(conn, host, deviceType):
    # Function that check if there is an entry on hosts file /etc/hosts for the device

    commandsReport = {}
    errors = {}
    i = 1
    localHostname = subprocess.run("hostname", shell=True, capture_output=True, text=True).stdout.strip("\n")
    device = host
    etcHosts = "/etc/hosts"
    client = SSHClient()
    client.set_missing_host_key_policy(AutoAddPolicy())
    command = "cat " + etcHosts + " | grep " + device
    command = "cat " + etcHosts + " | grep " + device
    # check if the script is running on dc13 or sv15 network-utils
    if localHostname == "network-utils01-dc13":
        jumpBox1 = ""
        jumpBox2 = "136.144.54.141"

    elif localHostname == "network-utils01-sv15":
        jumpBox1 = ""
        jumpBox2 = "136.144.56.77"

    else:
        jumpBox1 = "136.144.56.77"
        jumpBox2 = "136.144.54.141"
    # look for the device on /etc/hosts file
    check = subprocess.run(command, shell=True, stdout=subprocess.DEVNULL).returncode

    if check != 0:
        errors[i] = (
            "No entry found on: "
            + etcHosts
            + " ("
            + localHostname
            + ") for the device "
            + device
            + ", please review or add it manually."
        )
        commandsReport[i] = {1: command}
        i += 1

    client.connect(jumpBox2)
    stdin, stdout, stderr = client.exec_command(command)
    stdin2, stdout2, stderr2 = client.exec_command("hostname")
    remoteHostname = (stdout2.read().decode("utf8")).strip("\n")
    remoteCheck = stdout.read().decode("utf8")

    if jumpBox1:
        client.connect(jumpBox1)
        stdin, stdout, stderr = client.exec_command(command)
        stdin2, stdout2, stderr2 = client.exec_command("hostname")
        remoteHostname2 = (stdout2.read().decode("utf8")).strip("\n")
        remoteCheck2 = stdout.read().decode("utf8")

    if not remoteCheck:
        errors[i] = (
            "No entry found on: "
            + etcHosts
            + " ("
            + remoteHostname
            + ") for the device "
            + device
            + ", please review or add it manually."
        )
        commandsReport[i] = {1: command}
        i += 1

    if jumpBox1 and not remoteCheck2:
        errors[i] = (
            "No entry found on: "
            + etcHosts
            + " ("
            + remoteHostname2
            + ") for the device "
            + device
            + ", please review or add it manually."
        )
        commandsReport[i] = {1: command}
        i += 1

    if errors:
        printReport(
            conn,
            host,
            "Test 16 failed for " + host + " - There is no entry on /etc/hosts file for the device.",
            errors,
            commandsReport,
            "other",
        )
        return False
    else:
        printReport(
            conn,
            host,
            "Test 16 successfull for " + host + " - Found an entry for the device on /etc/hosts file.",
            "",
            "",
            "other",
        )
        return True
